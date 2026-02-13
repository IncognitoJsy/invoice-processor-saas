"""Quote Builder routes - electrical takeoff and estimation from drawings

This module now includes the interactive takeoff canvas functionality,
replacing the old automated AI parsing approach with a hybrid user-controlled workflow.
"""
from flask import Blueprint, render_template, jsonify, request, send_file, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime
from functools import wraps
from collections import defaultdict
import time
import os
import uuid
import json
import base64

PAPER_SIZES_MM = {
    'A0': (1189, 841),
    'A1': (841, 594),
    'A2': (594, 420),
    'A3': (420, 297),
    'A4': (297, 210),
}

bp = Blueprint('quotebuilder', __name__, url_prefix='/quotebuilder')


# Simple in-memory rate limiter
_rate_limit_store = defaultdict(list)


def rate_limit(max_calls, period_seconds):
    """Rate limit decorator. max_calls per period_seconds per user."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            user_id = current_user.id if current_user.is_authenticated else 'anon'
            key = f"{f.__name__}:{user_id}"
            now = time.time()
            
            # Clean old entries
            _rate_limit_store[key] = [t for t in _rate_limit_store[key] if now - t < period_seconds]
            
            if len(_rate_limit_store[key]) >= max_calls:
                return jsonify({'success': False, 'error': 'Too many requests. Please wait a moment.'}), 429
            
            _rate_limit_store[key].append(now)
            return f(*args, **kwargs)
        return wrapped
    return decorator


def _verify_project_ownership(project_id):
    """Verify the current user owns this project. Returns project or None."""
    from app.models.project import Project
    return Project.query.filter_by(id=project_id, user_id=current_user.id).first()


@bp.after_request
def add_security_headers(response):
    """Add security headers to all quotebuilder responses."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    # CSP - allow scripts from same origin, inline for Alpine.js, and CDNs
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: blob:; "
        "connect-src 'self'; "
        "frame-ancestors 'self'"
    )
    return response


# =============================================================================
# PAGE ROUTES
# =============================================================================

@bp.route('/')
@login_required
def index():
    """Quote Builder main page - list all projects"""
    return render_template('quotebuilder/index.html')


@bp.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    """Single project detail page"""
    from app.models.project import Project
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()
    return render_template('quotebuilder/project.html', project=project)


@bp.route('/project/new')
@login_required
def new_project():
    """Create new project page"""
    return render_template('quotebuilder/new_project.html')


@bp.route('/project/<int:project_id>/takeoff/<int:doc_id>')
@login_required
def takeoff_view(project_id, doc_id):
    """Open interactive takeoff canvas for a document"""
    from app.models.project import Project, ProjectDocument
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()
    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first_or_404()
    
    return render_template('quotebuilder/takeoff.html', project=project, document=document)


# =============================================================================
# PROJECT API ROUTES
# =============================================================================

@bp.route('/api/projects')
@login_required
def get_projects():
    """Get all projects for current user"""
    from app.models.project import Project
    
    projects = Project.query.filter_by(user_id=current_user.id)\
        .order_by(Project.updated_at.desc())\
        .all()
    
    # Stats
    total_quoted = sum(float(p.grand_total or 0) for p in projects if p.status in ['quoted', 'sent', 'won'])
    won_value = sum(float(p.grand_total or 0) for p in projects if p.status == 'won')
    
    return jsonify({
        'success': True,
        'projects': [p.to_dict() for p in projects],
        'stats': {
            'total_projects': len(projects),
            'draft': len([p for p in projects if p.status == 'draft']),
            'quoted': len([p for p in projects if p.status in ['quoted', 'sent']]),
            'won': len([p for p in projects if p.status == 'won']),
            'total_quoted_value': total_quoted,
            'won_value': won_value
        }
    })


@bp.route('/api/projects', methods=['POST'])
@login_required
@rate_limit(10, 60)
def create_project():
    """Create a new project"""
    from app.models.project import Project
    from app.extensions import db
    
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    # Validate and sanitise inputs
    name = str(data.get('name', 'New Project'))[:200].strip()
    if not name:
        name = 'New Project'
    
    client_name = str(data.get('client_name', '') or '')[:200].strip() or None
    client_email = str(data.get('client_email', '') or '')[:200].strip() or None
    client_phone = str(data.get('client_phone', '') or '')[:50].strip() or None
    site_address = str(data.get('site_address', '') or '')[:500].strip() or None
    
    # Validate numeric fields
    try:
        markup = min(max(float(data.get('materials_markup_percent', 25)), 0), 500)
    except (TypeError, ValueError):
        markup = 25
    try:
        labour_rate = min(max(float(data.get('labour_rate_per_hour', 45)), 0), 1000)
    except (TypeError, ValueError):
        labour_rate = 45
    try:
        contingency = min(max(float(data.get('contingency_percent', 10)), 0), 100)
    except (TypeError, ValueError):
        contingency = 10
    
    # Validate enum fields
    supply_type = data.get('supply_type', 'single_phase')
    if supply_type not in ('single_phase', 'three_phase'):
        supply_type = 'single_phase'
    building_type = data.get('building_type', 'renovation')
    if building_type not in ('new_build', 'renovation', 'retrofit', 'listed'):
        building_type = 'renovation'
    
    # Limit projects per user (max 100)
    project_count = Project.query.filter_by(user_id=current_user.id).count()
    if project_count >= 100:
        return jsonify({'success': False, 'error': 'Maximum 100 projects. Please delete old projects.'}), 400
    
    project = Project(
        user_id=current_user.id,
        name=name,
        client_name=client_name,
        client_email=client_email,
        client_phone=client_phone,
        site_address=site_address,
        supply_type=supply_type,
        building_type=building_type,
        materials_markup_percent=markup,
        labour_rate_per_hour=labour_rate,
        contingency_percent=contingency,
    )
    
    db.session.add(project)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'project': project.to_dict()
    })


@bp.route('/api/projects/<int:project_id>')
@login_required
def get_project(project_id):
    """Get single project with all details"""
    from app.models.project import Project, ProjectDocument, ProjectMaterial, ProjectLabour
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    documents = ProjectDocument.query.filter_by(project_id=project.id).all()
    materials = ProjectMaterial.query.filter_by(project_id=project.id).order_by(ProjectMaterial.category).all()
    labour = ProjectLabour.query.filter_by(project_id=project.id).all()
    
    # Group materials by category
    materials_by_category = {}
    for m in materials:
        cat = m.category or 'Uncategorised'
        if cat not in materials_by_category:
            materials_by_category[cat] = []
        materials_by_category[cat].append(m.to_dict())
    
    return jsonify({
        'success': True,
        'project': project.to_dict(),
        'documents': [d.to_dict() for d in documents],
        'materials': [m.to_dict() for m in materials],
        'materials_by_category': materials_by_category,
        'labour': [l.to_dict() for l in labour],
    })


@bp.route('/api/projects/<int:project_id>', methods=['PUT'])
@login_required
def update_project(project_id):
    """Update project settings"""
    from app.models.project import Project
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    # Validate and sanitise string fields with length limits
    string_limits = {
        'name': 200, 'client_name': 200, 'client_email': 200,
        'client_phone': 50, 'site_address': 500,
    }
    for field, max_len in string_limits.items():
        if field in data:
            val = str(data[field] or '')[:max_len].strip()
            setattr(project, field, val if val else None)
    
    # Validate enum fields
    if 'supply_type' in data:
        if data['supply_type'] in ('single_phase', 'three_phase'):
            project.supply_type = data['supply_type']
    if 'building_type' in data:
        if data['building_type'] in ('new_build', 'renovation', 'retrofit', 'listed'):
            project.building_type = data['building_type']
    if 'status' in data:
        if data['status'] in ('draft', 'quoted', 'sent', 'won', 'lost'):
            project.status = data['status']
    
    # Validate numeric fields with bounds
    if 'materials_markup_percent' in data:
        try:
            project.materials_markup_percent = min(max(float(data['materials_markup_percent']), 0), 500)
        except (TypeError, ValueError):
            pass
    if 'labour_rate_per_hour' in data:
        try:
            project.labour_rate_per_hour = min(max(float(data['labour_rate_per_hour']), 0), 1000)
        except (TypeError, ValueError):
            pass
    if 'contingency_percent' in data:
        try:
            project.contingency_percent = min(max(float(data['contingency_percent']), 0), 100)
        except (TypeError, ValueError):
            pass
    if 'quote_valid_days' in data:
        try:
            project.quote_valid_days = min(max(int(data['quote_valid_days']), 1), 365)
        except (TypeError, ValueError):
            pass
    
    # Recalculate if markup or contingency changed
    if 'materials_markup_percent' in data or 'contingency_percent' in data:
        # Update all material markups
        from app.models.project import ProjectMaterial
        if 'materials_markup_percent' in data:
            for material in ProjectMaterial.query.filter_by(project_id=project.id).all():
                material.calculate_totals(markup_percent=data['materials_markup_percent'])
        
        project.recalculate_totals()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'project': project.to_dict()
    })


@bp.route('/api/projects/<int:project_id>', methods=['DELETE'])
@login_required
def delete_project(project_id):
    """Delete a project and all related data"""
    from app.models.project import Project, ProjectDocument
    from app.models.takeoff import (
        TakeoffRoom, TakeoffSymbolTemplate, TakeoffSymbolDetection,
        TakeoffCableRun, TakeoffArea
    )
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    try:
        # Delete takeoff data first (foreign key constraints)
        TakeoffSymbolDetection.query.filter_by(project_id=project_id).delete()
        TakeoffSymbolTemplate.query.filter_by(project_id=project_id).delete()
        TakeoffCableRun.query.filter_by(project_id=project_id).delete()
        TakeoffRoom.query.filter_by(project_id=project_id).delete()
        TakeoffArea.query.filter_by(project_id=project_id).delete()
        
        # Now delete the project (cascades to documents, materials, etc.)
        db.session.delete(project)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting project {project_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to delete project. Please try again.'}), 500


# =============================================================================
# DOCUMENT UPLOAD & PARSING
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents', methods=['POST'])
@login_required
@rate_limit(10, 60)
def upload_document(project_id):
    """Upload drawings/specs to a project"""
    from app.models.project import Project, ProjectDocument
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Validate file type
    allowed_extensions = {'pdf', 'png', 'jpg', 'jpeg'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    
    if ext not in allowed_extensions:
        return jsonify({'success': False, 'error': f'File type .{ext} not allowed. Accepted: PDF, PNG, JPG'}), 400
    
    # Validate file size (max 50MB)
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    max_size = 50 * 1024 * 1024  # 50MB
    if file_size > max_size:
        return jsonify({'success': False, 'error': f'File too large ({file_size // (1024*1024)}MB). Maximum is 50MB.'}), 400
    
    # Validate MIME type matches extension
    allowed_mimes = {
        'pdf': ['application/pdf'],
        'png': ['image/png'],
        'jpg': ['image/jpeg'],
        'jpeg': ['image/jpeg'],
    }
    if file.content_type not in allowed_mimes.get(ext, []):
        current_app.logger.warning(f"MIME mismatch: {file.content_type} for .{ext} from user {current_user.id}")
    
    # Limit documents per project (max 20)
    existing_count = ProjectDocument.query.filter_by(project_id=project_id).count()
    if existing_count >= 20:
        return jsonify({'success': False, 'error': 'Maximum 20 documents per project'}), 400
    
    # Save file
    filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
    upload_folder = os.path.join(current_app.root_path, 'uploads', 'projects', str(project_id))
    os.makedirs(upload_folder, exist_ok=True)
    
    file_path = os.path.join(upload_folder, filename)
    file.save(file_path)
    
    # Create document record
    document = ProjectDocument(
        project_id=project.id,
        filename=filename,
        original_filename=file.filename,
        file_path=file_path,
        file_size=os.path.getsize(file_path),
        mime_type=file.content_type,
        document_type=request.form.get('document_type', 'drawing'),
        floor_level=request.form.get('floor_level'),
        system_type=request.form.get('system_type', 'all'),
    )
    
    db.session.add(document)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'document': document.to_dict()
    })


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/parse', methods=['POST'])
@login_required
@rate_limit(3, 60)
def parse_document(project_id, doc_id):
    """Parse a drawing using AI to extract materials (legacy - now use takeoff canvas instead)"""
    from app.models.project import Project, ProjectDocument, ProjectMaterial
    from app.extensions import db
    from app.parsers.drawing_parser import DrawingParser
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404
    
    try:
        # Parse the drawing
        parser = DrawingParser()
        result = parser.parse(
            document.file_path,
            document_type=document.document_type,
            system_type=document.system_type,
            floor_level=document.floor_level
        )
        
        if not result.get('success'):
            document.parse_error = result.get('error', 'Unknown parsing error')
            db.session.commit()
            return jsonify({'success': False, 'error': result.get('error')}), 400
        
        # Add extracted materials to project
        materials_added = 0
        for item in result.get('materials', []):
            # Check if similar material already exists
            existing = ProjectMaterial.query.filter_by(
                project_id=project.id,
                part_number=item.get('part_number'),
                category=item.get('category')
            ).first()
            
            if existing:
                # Update quantity
                existing.quantity = float(existing.quantity or 0) + float(item.get('quantity', 0))
                existing.calculate_totals(markup_percent=float(project.materials_markup_percent))
            else:
                # Add new material
                material = ProjectMaterial(
                    project_id=project.id,
                    source_document_id=document.id,
                    category=item.get('category'),
                    part_number=item.get('part_number'),
                    description=item.get('description'),
                    manufacturer=item.get('manufacturer'),
                    quantity=item.get('quantity', 1),
                    unit=item.get('unit', 'each'),
                    unit_cost=item.get('unit_cost'),
                    price_source=item.get('price_source', 'estimated'),
                )
                material.calculate_totals(markup_percent=float(project.materials_markup_percent))
                db.session.add(material)
                materials_added += 1
        
        # Update document status
        document.parsed = True
        document.parsed_at = datetime.utcnow()
        document.scale = result.get('scale')
        document.drawing_number = result.get('drawing_number')
        
        # Recalculate project totals
        project.recalculate_totals()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'materials_added': materials_added,
            'materials_updated': len(result.get('materials', [])) - materials_added,
            'document': document.to_dict()
        })
        
    except Exception as e:
        current_app.logger.error(f"Drawing parse error: {str(e)}")
        document.parse_error = str(e)
        db.session.commit()
        return jsonify({'success': False, 'error': 'Failed to parse document. Please check the file and try again.'}), 500


# =============================================================================
# MATERIALS API
# =============================================================================

@bp.route('/api/projects/<int:project_id>/materials', methods=['POST'])
@login_required
def add_material(project_id):
    """Manually add a material to project"""
    from app.models.project import Project, ProjectMaterial
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    # Validate quantity
    try:
        quantity = max(float(data.get('quantity', 1)), 0.01)
    except (TypeError, ValueError):
        quantity = 1
    
    # Validate unit cost
    try:
        unit_cost = max(float(data.get('unit_cost', 0)), 0)
    except (TypeError, ValueError):
        unit_cost = 0
    
    material = ProjectMaterial(
        project_id=project.id,
        manually_added=True,
        category=str(data.get('category', 'Uncategorised') or 'Uncategorised')[:100].strip(),
        part_number=str(data.get('part_number', '') or '')[:100].strip() or None,
        description=str(data.get('description', '') or '')[:500].strip() or None,
        manufacturer=str(data.get('manufacturer', '') or '')[:200].strip() or None,
        quantity=quantity,
        unit=str(data.get('unit', 'each') or 'each')[:20].strip(),
        unit_cost=unit_cost,
        price_source='manual',
        notes=str(data.get('notes', '') or '')[:1000].strip() or None,
    )
    
    material.calculate_totals(markup_percent=float(project.materials_markup_percent))
    
    db.session.add(material)
    project.recalculate_totals()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'material': material.to_dict(),
        'project_totals': {
            'total_materials_cost': float(project.total_materials_cost),
            'total_materials_sell': float(project.total_materials_sell),
            'grand_total': float(project.grand_total),
        }
    })


@bp.route('/api/projects/<int:project_id>/materials/<int:material_id>', methods=['PUT'])
@login_required
def update_material(project_id, material_id):
    """Update a material"""
    from app.models.project import Project, ProjectMaterial
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    material = ProjectMaterial.query.filter_by(id=material_id, project_id=project_id).first()
    if not material:
        return jsonify({'success': False, 'error': 'Material not found'}), 404
    
    data = request.get_json()
    
    for field in ['category', 'part_number', 'description', 'manufacturer',
                  'quantity', 'unit', 'unit_cost', 'markup_percent', 'notes']:
        if field in data:
            setattr(material, field, data[field])
    
    material.calculate_totals()
    project.recalculate_totals()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'material': material.to_dict(),
        'project_totals': {
            'total_materials_cost': float(project.total_materials_cost),
            'total_materials_sell': float(project.total_materials_sell),
            'grand_total': float(project.grand_total),
        }
    })


@bp.route('/api/projects/<int:project_id>/materials/<int:material_id>', methods=['DELETE'])
@login_required
def delete_material(project_id, material_id):
    """Delete a material"""
    from app.models.project import Project, ProjectMaterial
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    material = ProjectMaterial.query.filter_by(id=material_id, project_id=project_id).first()
    if not material:
        return jsonify({'success': False, 'error': 'Material not found'}), 404
    
    db.session.delete(material)
    project.recalculate_totals()
    db.session.commit()
    
    return jsonify({'success': True})


# =============================================================================
# PRICING - QB/XERO MATCHING
# =============================================================================

@bp.route('/api/projects/<int:project_id>/match-prices', methods=['POST'])
@login_required
@rate_limit(5, 60)
def match_prices_from_accounting(project_id):
    """Match materials to QuickBooks/Xero products and pull prices"""
    from app.models.project import Project, ProjectMaterial
    from app.models.quickbooks import QuickBooksConnection
    from app.integrations.quickbooks_service import QuickBooksService
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    # Get QB connection
    qb_connection = QuickBooksConnection.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).first()
    
    if not qb_connection:
        return jsonify({'success': False, 'error': 'QuickBooks not connected'}), 400
    
    qb_service = QuickBooksService()
    
    # Get all QB items
    response = qb_service.get_items(qb_connection)
    if 'error' in response:
        return jsonify({'success': False, 'error': response['error']}), 400
    
    qb_items = response.get('QueryResponse', {}).get('Item', [])
    
    # Build lookup by SKU
    qb_by_sku = {}
    for item in qb_items:
        sku = item.get('Sku', '').upper()
        if sku:
            qb_by_sku[sku] = item
    
    # Match materials
    materials = ProjectMaterial.query.filter_by(project_id=project_id).all()
    matched = 0
    unmatched = []
    
    for material in materials:
        part_upper = (material.part_number or '').upper()
        
        if part_upper in qb_by_sku:
            qb_item = qb_by_sku[part_upper]
            
            # Get purchase cost and sales price
            purchase_cost = float(qb_item.get('PurchaseCost', 0) or 0)
            sales_price = float(qb_item.get('UnitPrice', 0) or 0)
            
            material.unit_cost = purchase_cost if purchase_cost > 0 else material.unit_cost
            material.qb_item_id = qb_item.get('Id')
            material.qb_item_name = qb_item.get('Name')
            material.price_source = 'quickbooks'
            material.price_verified = True
            material.price_date = datetime.utcnow()
            
            # Use higher of QB sales price or calculated price
            material.calculate_totals(markup_percent=float(project.materials_markup_percent))
            if sales_price > float(material.unit_sell or 0):
                material.unit_sell = sales_price
                material.total_sell = round(float(material.quantity) * sales_price, 2)
            
            matched += 1
        else:
            unmatched.append({
                'part_number': material.part_number,
                'description': material.description
            })
    
    project.recalculate_totals()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'matched': matched,
        'unmatched_count': len(unmatched),
        'unmatched': unmatched[:20],  # Return first 20 unmatched
        'project_totals': {
            'total_materials_cost': float(project.total_materials_cost),
            'total_materials_sell': float(project.total_materials_sell),
            'grand_total': float(project.grand_total),
        }
    })


# =============================================================================
# SUPPLIER QUOTE REQUESTS
# =============================================================================

@bp.route('/api/projects/<int:project_id>/supplier-request', methods=['POST'])
@login_required
def generate_supplier_request(project_id):
    """Generate supplier quote request spreadsheet"""
    from app.models.project import Project, ProjectMaterial, SupplierQuoteRequest
    from app.extensions import db
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    data = request.get_json()
    category_filter = data.get('category')  # Optional - filter by category
    supplier_name = data.get('supplier_name', 'Supplier')
    
    # Get materials
    query = ProjectMaterial.query.filter_by(project_id=project_id)
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    materials = query.filter(
        (ProjectMaterial.price_verified == False) | (ProjectMaterial.price_source == 'estimated')
    ).all()
    
    if not materials:
        return jsonify({'success': False, 'error': 'No materials need pricing'}), 400
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Request"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    
    # Header info
    ws['A1'] = f"QUOTATION REQUEST - {project.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Project: {project.site_address or project.name}"
    ws['A3'] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A4'] = f"Contact: {current_user.company_name or current_user.email}"
    
    # Column headers
    headers = ["Item", "Description", "Manufacturer", "Part Number", "Qty", "Unit", "Your Price", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    for idx, material in enumerate(materials, 1):
        row = 6 + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=material.description)
        ws.cell(row=row, column=3, value=material.manufacturer)
        ws.cell(row=row, column=4, value=material.part_number)
        ws.cell(row=row, column=5, value=float(material.quantity))
        ws.cell(row=row, column=6, value=material.unit)
        ws.cell(row=row, column=7, value="")  # Supplier fills in
        ws.cell(row=row, column=8, value=f"=E{row}*G{row}")
    
    # Column widths
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Track the request
    quote_request = SupplierQuoteRequest(
        project_id=project.id,
        supplier_name=supplier_name,
        category=category_filter,
        items_count=len(materials),
        status='pending'
    )
    db.session.add(quote_request)
    db.session.commit()
    
    # Return file
    filename = f"{project.name.replace(' ', '_')}_Quote_Request_{supplier_name}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =============================================================================
# LABOUR CALCULATIONS
# =============================================================================

@bp.route('/api/projects/<int:project_id>/calculate-labour', methods=['POST'])
@login_required
def calculate_labour(project_id):
    """Auto-calculate labour based on materials quantities"""
    from app.models.project import Project, ProjectMaterial, ProjectLabour
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    # Delete existing auto-calculated labour
    ProjectLabour.query.filter_by(project_id=project_id, auto_calculated=True).delete()
    
    # Get materials
    materials = ProjectMaterial.query.filter_by(project_id=project_id).all()
    
    # Labour rates per task (hours per unit)
    labour_rates = {
        # First fix times (per point)
        'downlights': {'first_fix': 0.25, 'second_fix': 0.15},
        'sockets': {'first_fix': 0.5, 'second_fix': 0.17},
        'switches': {'first_fix': 0.25, 'second_fix': 0.17},
        'fused_spurs': {'first_fix': 0.5, 'second_fix': 0.17},
        'data_points': {'first_fix': 0.5, 'second_fix': 0.25},
        'smoke_detectors': {'first_fix': 0.25, 'second_fix': 0.17},
        'distribution': {'install': 8},  # Per board
        'cable_per_100m': {'install': 2},
        'led_tape_per_m': {'install': 0.5},
        'containment_per_m': {'install': 0.1},
        'testing': {'per_circuit': 0.25},
    }
    
    # Building type multiplier
    multipliers = {
        'new_build': 0.7,
        'renovation': 1.0,
        'retrofit': 1.5,
        'listed': 2.0,
    }
    multiplier = multipliers.get(project.building_type, 1.0)
    
    # Count materials by type
    counts = {
        'downlights': 0,
        'sockets': 0,
        'switches': 0,
        'fused_spurs': 0,
        'data_points': 0,
        'smoke_detectors': 0,
        'distribution_boards': 0,
        'cable_metres': 0,
        'led_tape_metres': 0,
        'containment_metres': 0,
    }
    
    for m in materials:
        cat = (m.category or '').lower()
        desc = (m.description or '').lower()
        qty = float(m.quantity or 0)
        
        if 'downlight' in desc or 'light_type_a' in (m.part_number or '').lower():
            counts['downlights'] += qty
        elif 'socket' in desc and 'double' in desc:
            counts['sockets'] += qty
        elif 'switch' in desc or 'dimmer' in desc:
            counts['switches'] += qty
        elif 'spur' in desc or 'fcu' in desc:
            counts['fused_spurs'] += qty
        elif 'data' in desc or 'cat6' in desc.lower():
            counts['data_points'] += qty
        elif 'smoke' in desc or 'heat' in desc or 'detector' in desc:
            counts['smoke_detectors'] += qty
        elif 'consumer unit' in desc or 'distribution' in desc:
            counts['distribution_boards'] += qty
        elif 'cable' in cat and m.unit == 'm':
            counts['cable_metres'] += qty
        elif 'led tape' in desc or 'tape' in cat:
            counts['led_tape_metres'] += qty
        elif 'conduit' in desc or 'trunking' in desc:
            counts['containment_metres'] += qty
    
    # Generate labour items
    labour_items = []
    rate = float(project.labour_rate_per_hour)
    
    # First fix
    first_fix_hours = (
        counts['downlights'] * labour_rates['downlights']['first_fix'] +
        counts['sockets'] * labour_rates['sockets']['first_fix'] +
        counts['switches'] * labour_rates['switches']['first_fix'] +
        counts['fused_spurs'] * labour_rates['fused_spurs']['first_fix'] +
        counts['data_points'] * labour_rates['data_points']['first_fix'] +
        counts['smoke_detectors'] * labour_rates['smoke_detectors']['first_fix']
    ) * multiplier
    
    if first_fix_hours > 0:
        labour = ProjectLabour(
            project_id=project.id,
            task="First Fix (cabling & back boxes)",
            hours=round(first_fix_hours, 1),
            rate=rate,
            auto_calculated=True,
            calculation_basis=f"{int(counts['sockets'])} sockets, {int(counts['switches'])} switches, {int(counts['downlights'])} lights"
        )
        labour.calculate_total()
        db.session.add(labour)
        labour_items.append(labour.to_dict())
    
    # Second fix
    second_fix_hours = (
        counts['downlights'] * labour_rates['downlights']['second_fix'] +
        counts['sockets'] * labour_rates['sockets']['second_fix'] +
        counts['switches'] * labour_rates['switches']['second_fix'] +
        counts['fused_spurs'] * labour_rates['fused_spurs']['second_fix'] +
        counts['data_points'] * labour_rates['data_points']['second_fix'] +
        counts['smoke_detectors'] * labour_rates['smoke_detectors']['second_fix']
    ) * multiplier
    
    if second_fix_hours > 0:
        labour = ProjectLabour(
            project_id=project.id,
            task="Second Fix (accessories & fittings)",
            hours=round(second_fix_hours, 1),
            rate=rate,
            auto_calculated=True,
        )
        labour.calculate_total()
        db.session.add(labour)
        labour_items.append(labour.to_dict())
    
    # Distribution boards
    if counts['distribution_boards'] > 0:
        db_hours = counts['distribution_boards'] * labour_rates['distribution']['install'] * multiplier
        labour = ProjectLabour(
            project_id=project.id,
            task="Distribution Board Installation",
            hours=round(db_hours, 1),
            rate=rate,
            auto_calculated=True,
            calculation_basis=f"{int(counts['distribution_boards'])} board(s)"
        )
        labour.calculate_total()
        db.session.add(labour)
        labour_items.append(labour.to_dict())
    
    # LED tape
    if counts['led_tape_metres'] > 0:
        tape_hours = counts['led_tape_metres'] * labour_rates['led_tape_per_m']['install']
        labour = ProjectLabour(
            project_id=project.id,
            task="LED Tape Installation",
            hours=round(tape_hours, 1),
            rate=rate,
            auto_calculated=True,
            calculation_basis=f"{int(counts['led_tape_metres'])}m of LED tape"
        )
        labour.calculate_total()
        db.session.add(labour)
        labour_items.append(labour.to_dict())
    
    # Containment
    if counts['containment_metres'] > 0:
        cont_hours = counts['containment_metres'] * labour_rates['containment_per_m']['install'] * multiplier
        labour = ProjectLabour(
            project_id=project.id,
            task="Containment Installation",
            hours=round(cont_hours, 1),
            rate=rate,
            auto_calculated=True,
        )
        labour.calculate_total()
        db.session.add(labour)
        labour_items.append(labour.to_dict())
    
    # Testing (estimate circuits from DB count × 10 ways average)
    estimated_circuits = max(10, counts['distribution_boards'] * 10)
    test_hours = estimated_circuits * labour_rates['testing']['per_circuit']
    labour = ProjectLabour(
        project_id=project.id,
        task="Testing & Certification",
        hours=round(test_hours, 1),
        rate=rate,
        auto_calculated=True,
        calculation_basis=f"~{estimated_circuits} circuits"
    )
    labour.calculate_total()
    db.session.add(labour)
    labour_items.append(labour.to_dict())
    
    # Recalculate totals
    project.recalculate_totals()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'labour': labour_items,
        'total_hours': float(project.total_labour_hours),
        'total_cost': float(project.total_labour_cost),
        'project_totals': {
            'subtotal': float(project.subtotal),
            'contingency': float(project.contingency_amount),
            'grand_total': float(project.grand_total),
        }
    })


# =============================================================================
# QUOTE GENERATION
# =============================================================================

@bp.route('/api/projects/<int:project_id>/generate-quote', methods=['POST'])
@login_required
@rate_limit(10, 60)
def generate_quote_pdf(project_id):
    """Generate professional PDF quotation"""
    from app.models.project import Project, ProjectMaterial, ProjectLabour
    from app.extensions import db
    from datetime import datetime, timedelta
    import os
    import io

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    materials = ProjectMaterial.query.filter_by(project_id=project.id).order_by(ProjectMaterial.category).all()
    labour_items = ProjectLabour.query.filter_by(project_id=project.id).all()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.pdfgen import canvas
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        # Generate filename
        safe_name = (project.name or 'Quote').replace(' ', '_').replace('/', '-')
        filename = f"Quote_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        upload_dir = os.path.join(current_app.root_path, '..', 'uploads', 'quotes')
        os.makedirs(upload_dir, exist_ok=True)
        filepath = os.path.join(upload_dir, filename)

        width, height = A4
        c = canvas.Canvas(filepath, pagesize=A4)

        # Colours
        primary = HexColor('#4F46E5')     # Indigo
        dark = HexColor('#1F2937')
        grey = HexColor('#6B7280')
        light_grey = HexColor('#F3F4F6')
        green = HexColor('#059669')
        line_col = HexColor('#E5E7EB')

        page_num = 1
        margin_left = 30 * mm
        margin_right = width - 30 * mm
        content_width = margin_right - margin_left

        def draw_header(c, y):
            """Draw page header with company branding"""
            # Logo / Company name
            c.setFont('Helvetica-Bold', 22)
            c.setFillColor(primary)
            c.drawString(margin_left, y, 'GoZappify')
            c.setFont('Helvetica', 8)
            c.setFillColor(grey)
            c.drawString(margin_left, y - 14, 'Electrical Quotation')

            # Quote reference top-right
            c.setFont('Helvetica-Bold', 10)
            c.setFillColor(dark)
            c.drawRightString(margin_right, y, f'QUOTE #{project.id:04d}')
            c.setFont('Helvetica', 9)
            c.setFillColor(grey)
            c.drawRightString(margin_right, y - 13, datetime.utcnow().strftime('%d %B %Y'))
            return y - 35

        def draw_footer(c, page_num):
            """Draw page footer"""
            c.setFont('Helvetica', 7)
            c.setFillColor(grey)
            c.drawString(margin_left, 15 * mm, f'Quote #{project.id:04d} — {project.name or ""}')
            c.drawRightString(margin_right, 15 * mm, f'Page {page_num}')
            # Line above footer
            c.setStrokeColor(line_col)
            c.setLineWidth(0.5)
            c.line(margin_left, 20 * mm, margin_right, 20 * mm)

        def new_page(c, page_num):
            c.showPage()
            page_num += 1
            y = height - 25 * mm
            y = draw_header(c, y)
            return y, page_num

        def check_space(c, y, needed, page_num):
            if y - needed < 30 * mm:
                draw_footer(c, page_num)
                y, page_num = new_page(c, page_num)
            return y, page_num

        # ── PAGE 1: Header + Client/Project Details ──────────────
        y = height - 25 * mm
        y = draw_header(c, y)

        y -= 10

        # QUOTATION title bar
        c.setFillColor(primary)
        c.roundRect(margin_left, y - 28, content_width, 30, 4, fill=1, stroke=0)
        c.setFont('Helvetica-Bold', 14)
        c.setFillColor(white)
        c.drawString(margin_left + 10, y - 20, 'ELECTRICAL QUOTATION')
        y -= 45

        # Two-column: Client details | Project details
        col_w = content_width / 2 - 5

        # Left column - Client
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(primary)
        c.drawString(margin_left, y, 'CLIENT')
        c.setFont('Helvetica', 9)
        c.setFillColor(dark)
        y -= 14
        client_name = project.client_name or 'TBC'
        c.drawString(margin_left, y, client_name)
        y -= 12
        if project.client_email:
            c.drawString(margin_left, y, project.client_email)
            y -= 12
        if project.client_phone:
            c.drawString(margin_left, y, project.client_phone)
            y -= 12

        # Right column - Project
        right_x = margin_left + col_w + 10
        py = y + 14 + 12 + (12 if project.client_email else 0) + (12 if project.client_phone else 0)
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(primary)
        c.drawString(right_x, py, 'PROJECT')
        c.setFont('Helvetica', 9)
        c.setFillColor(dark)
        py -= 14
        c.drawString(right_x, py, project.name or 'Untitled')
        py -= 12
        if project.site_address:
            c.drawString(right_x, py, project.site_address)
            py -= 12
        valid_days = project.quote_valid_days or 30
        valid_until = (datetime.utcnow() + timedelta(days=valid_days)).strftime('%d %B %Y')
        c.setFont('Helvetica', 8)
        c.setFillColor(grey)
        c.drawString(right_x, py, f'Valid until: {valid_until}')

        y -= 20

        # Divider
        c.setStrokeColor(line_col)
        c.setLineWidth(0.5)
        c.line(margin_left, y, margin_right, y)
        y -= 15

        # ── MATERIALS TABLE ──────────────────────────────────────
        # Group by category
        materials_by_cat = {}
        for m in materials:
            cat = m.category or 'General'
            if cat not in materials_by_cat:
                materials_by_cat[cat] = []
            materials_by_cat[cat].append(m)

        if materials:
            c.setFont('Helvetica-Bold', 11)
            c.setFillColor(dark)
            c.drawString(margin_left, y, 'Materials')
            y -= 5

            # Table header
            y -= 15
            c.setFillColor(light_grey)
            c.rect(margin_left, y - 3, content_width, 16, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5)
            c.setFillColor(grey)
            c.drawString(margin_left + 4, y, 'Item')
            c.drawString(margin_left + content_width * 0.50, y, 'Qty')
            c.drawRightString(margin_left + content_width * 0.70, y, 'Unit Price')
            c.drawRightString(margin_right - 4, y, 'Total')
            y -= 18

            total_materials_sell = 0

            for cat, cat_materials in materials_by_cat.items():
                y, page_num = check_space(c, y, 25, page_num)

                # Category header
                c.setFont('Helvetica-Bold', 8)
                c.setFillColor(primary)
                c.drawString(margin_left + 4, y, cat)
                y -= 14

                for m in cat_materials:
                    y, page_num = check_space(c, y, 14, page_num)

                    unit_sell = float(m.unit_sell or 0)
                    qty = int(m.quantity or 0)
                    line_total = float(m.total_sell or 0)
                    total_materials_sell += line_total

                    c.setFont('Helvetica', 8)
                    c.setFillColor(dark)

                    # Truncate description if too long
                    desc = (m.description or m.part_number or 'Item')[:55]
                    sku_text = f"({m.part_number}) " if m.part_number else ''
                    c.drawString(margin_left + 8, y, f'{sku_text}{desc}')
                    c.drawString(margin_left + content_width * 0.50, y, str(qty))
                    c.drawRightString(margin_left + content_width * 0.70, y, f'£{unit_sell:,.2f}')
                    c.drawRightString(margin_right - 4, y, f'£{line_total:,.2f}')

                    # Light line under each row
                    y -= 3
                    c.setStrokeColor(line_col)
                    c.setLineWidth(0.3)
                    c.line(margin_left, y, margin_right, y)
                    y -= 11

            # Materials subtotal
            y -= 5
            y, page_num = check_space(c, y, 20, page_num)
            c.setFont('Helvetica-Bold', 9)
            c.setFillColor(dark)
            c.drawString(margin_left + 8, y, 'Materials Subtotal')
            c.drawRightString(margin_right - 4, y, f'£{total_materials_sell:,.2f}')
            y -= 20

        # ── LABOUR ───────────────────────────────────────────────
        total_labour = float(project.total_labour_cost or 0)
        total_hours = float(project.total_labour_hours or 0)

        if total_labour > 0:
            y, page_num = check_space(c, y, 40, page_num)
            c.setFont('Helvetica-Bold', 11)
            c.setFillColor(dark)
            c.drawString(margin_left, y, 'Labour')
            y -= 18

            labour_rate = float(project.labour_rate_per_hour or 0)
            c.setFont('Helvetica', 9)
            c.setFillColor(dark)
            c.drawString(margin_left + 8, y, f'{total_hours:.1f} hours @ £{labour_rate:.2f}/hr')
            c.drawRightString(margin_right - 4, y, f'£{total_labour:,.2f}')
            y -= 20

        # ── TOTALS BOX ───────────────────────────────────────────
        y -= 10
        y, page_num = check_space(c, y, 80, page_num)

        # Totals background
        box_height = 70
        c.setFillColor(HexColor('#F9FAFB'))
        c.setStrokeColor(line_col)
        c.setLineWidth(1)
        c.roundRect(margin_left, y - box_height + 15, content_width, box_height, 6, fill=1, stroke=1)

        ty = y + 5
        right_col = margin_right - 10

        # Materials
        mat_sell = float(project.total_materials_sell or 0)
        c.setFont('Helvetica', 9)
        c.setFillColor(grey)
        c.drawString(margin_left + 10, ty, 'Materials')
        c.setFillColor(dark)
        c.drawRightString(right_col, ty, f'£{mat_sell:,.2f}')
        ty -= 14

        # Labour
        c.setFillColor(grey)
        c.drawString(margin_left + 10, ty, 'Labour')
        c.setFillColor(dark)
        c.drawRightString(right_col, ty, f'£{total_labour:,.2f}')
        ty -= 14

        # Contingency
        contingency = float(project.contingency_amount or 0)
        cont_pct = float(project.contingency_percent or 0)
        c.setFillColor(grey)
        c.drawString(margin_left + 10, ty, f'Contingency ({cont_pct:.0f}%)')
        c.setFillColor(dark)
        c.drawRightString(right_col, ty, f'£{contingency:,.2f}')
        ty -= 6

        # Divider line
        c.setStrokeColor(primary)
        c.setLineWidth(1)
        c.line(margin_left + 10, ty, margin_right - 10, ty)
        ty -= 16

        # Grand total
        grand_total = float(project.grand_total or 0)
        c.setFont('Helvetica-Bold', 13)
        c.setFillColor(dark)
        c.drawString(margin_left + 10, ty, 'TOTAL (excl. VAT)')
        c.setFillColor(primary)
        c.drawRightString(right_col, ty, f'£{grand_total:,.2f}')

        y = ty - 30

        # ── TERMS ────────────────────────────────────────────────
        y, page_num = check_space(c, y, 80, page_num)

        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(dark)
        c.drawString(margin_left, y, 'Terms & Conditions')
        y -= 14

        terms = [
            f'This quotation is valid for {valid_days} days from the date of issue.',
            'All prices are exclusive of VAT which will be charged at the prevailing rate.',
            'Payment terms: 30 days from date of invoice.',
            'Any additional works not specified in this quotation will be charged accordingly.',
            'All electrical work will be carried out in accordance with BS 7671 (18th Edition).',
            'An Electrical Installation Certificate will be provided upon completion.',
        ]

        c.setFont('Helvetica', 7.5)
        c.setFillColor(grey)
        for term in terms:
            y, page_num = check_space(c, y, 12, page_num)
            c.drawString(margin_left + 4, y, f'•  {term}')
            y -= 11

        # Footer
        draw_footer(c, page_num)

        c.save()

        # Update project status
        project.status = 'quoted'
        project.quoted_at = datetime.utcnow()
        db.session.commit()

        current_app.logger.info(f"Quote PDF generated: {filepath}")

        # Return the PDF as a download
        from flask import send_file
        return send_file(
            filepath,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Quote PDF generation error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to generate quote PDF. Please try again.'}), 500


@bp.route('/api/projects/<int:project_id>/generate-materials-list', methods=['POST'])
@login_required
@rate_limit(10, 60)
def generate_materials_list(project_id):
    """Generate internal materials list PDF with cost, sell and profit breakdown"""
    from app.models.project import Project, ProjectMaterial, ProjectLabour
    from app.extensions import db
    from datetime import datetime
    import os

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    materials = ProjectMaterial.query.filter_by(project_id=project.id).order_by(ProjectMaterial.category).all()

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.pdfgen import canvas

        safe_name = (project.name or 'Materials').replace(' ', '_').replace('/', '-')
        filename = f"Materials_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        upload_dir = os.path.join(current_app.root_path, '..', 'uploads', 'quotes')
        os.makedirs(upload_dir, exist_ok=True)
        filepath = os.path.join(upload_dir, filename)

        # Landscape for wider table
        width, height = landscape(A4)
        c = canvas.Canvas(filepath, pagesize=landscape(A4))

        primary = HexColor('#4F46E5')
        dark = HexColor('#1F2937')
        grey = HexColor('#6B7280')
        light_grey = HexColor('#F3F4F6')
        green = HexColor('#059669')
        red = HexColor('#DC2626')
        line_col = HexColor('#E5E7EB')

        margin_left = 20 * mm
        margin_right = width - 20 * mm
        content_width = margin_right - margin_left
        page_num = 1

        def draw_header(c, y):
            c.setFont('Helvetica-Bold', 18)
            c.setFillColor(primary)
            c.drawString(margin_left, y, 'GoZappify')
            c.setFont('Helvetica', 8)
            c.setFillColor(grey)
            c.drawString(margin_left, y - 13, 'INTERNAL — Materials & Profit Tracker')

            c.setFont('Helvetica-Bold', 10)
            c.setFillColor(dark)
            c.drawRightString(margin_right, y, project.name or 'Untitled')
            c.setFont('Helvetica', 8)
            c.setFillColor(grey)
            c.drawRightString(margin_right, y - 13, f'Generated: {datetime.utcnow().strftime("%d %b %Y")}')
            if project.client_name:
                c.drawRightString(margin_right, y - 24, f'Client: {project.client_name}')
            return y - 40

        def draw_footer(c, page_num):
            c.setFont('Helvetica', 7)
            c.setFillColor(grey)
            c.drawString(margin_left, 12 * mm, f'CONFIDENTIAL — {project.name} — Materials List')
            c.drawRightString(margin_right, 12 * mm, f'Page {page_num}')
            c.setFont('Helvetica-Bold', 7)
            c.setFillColor(red)
            c.drawCentredString(width / 2, 12 * mm, 'INTERNAL USE ONLY — DO NOT SHARE WITH CLIENT')
            c.setStrokeColor(line_col)
            c.setLineWidth(0.5)
            c.line(margin_left, 17 * mm, margin_right, 17 * mm)

        def new_page(c, page_num):
            c.showPage()
            page_num += 1
            y = height - 18 * mm
            y = draw_header(c, y)
            draw_table_header(c, y)
            return y - 16, page_num

        # Column positions (landscape A4 = ~297mm wide, ~190mm usable)
        cols = {
            'part':   margin_left + 4,
            'desc':   margin_left + content_width * 0.12,
            'qty':    margin_left + content_width * 0.48,
            'cost_u': margin_left + content_width * 0.55,
            'cost_t': margin_left + content_width * 0.65,
            'sell_u': margin_left + content_width * 0.74,
            'sell_t': margin_left + content_width * 0.83,
            'profit': margin_left + content_width * 0.92,
            'margin': margin_right - 4,
        }

        def draw_table_header(c, y):
            c.setFillColor(dark)
            c.rect(margin_left, y - 3, content_width, 16, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7)
            c.setFillColor(white)
            c.drawString(cols['part'], y, 'Part No.')
            c.drawString(cols['desc'], y, 'Description')
            c.drawString(cols['qty'], y, 'Qty')
            c.drawRightString(cols['cost_t'] - 4, y, 'Unit Cost')
            c.drawRightString(cols['sell_u'] - 4, y, 'Total Cost')
            c.drawRightString(cols['sell_t'] - 4, y, 'Unit Sell')
            c.drawRightString(cols['profit'] - 4, y, 'Total Sell')
            c.drawRightString(cols['margin'], y, 'Profit    %')

        def check_space(c, y, needed, page_num):
            if y - needed < 25 * mm:
                draw_footer(c, page_num)
                y, page_num = new_page(c, page_num)
            return y, page_num

        # ── PAGE 1 ───────────────────────────────────────────────
        y = height - 18 * mm
        y = draw_header(c, y)

        # CONFIDENTIAL banner
        c.setFillColor(HexColor('#FEF2F2'))
        c.setStrokeColor(red)
        c.setLineWidth(1)
        c.roundRect(margin_left, y - 14, content_width, 18, 3, fill=1, stroke=1)
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(red)
        c.drawCentredString(width / 2, y - 9, 'CONFIDENTIAL — INTERNAL COST & PROFIT BREAKDOWN — NOT FOR CLIENT')
        y -= 25

        # Table header
        draw_table_header(c, y)
        y -= 16

        # Group materials
        materials_by_cat = {}
        for m in materials:
            cat = m.category or 'General'
            if cat not in materials_by_cat:
                materials_by_cat[cat] = []
            materials_by_cat[cat].append(m)

        grand_cost = 0
        grand_sell = 0

        for cat, cat_materials in materials_by_cat.items():
            y, page_num = check_space(c, y, 20, page_num)

            # Category header
            c.setFillColor(HexColor('#EEF2FF'))
            c.rect(margin_left, y - 3, content_width, 14, fill=1, stroke=0)
            c.setFont('Helvetica-Bold', 7.5)
            c.setFillColor(primary)
            c.drawString(margin_left + 4, y, cat)

            cat_cost = sum(float(m.total_cost or 0) for m in cat_materials)
            cat_sell = sum(float(m.total_sell or 0) for m in cat_materials)
            cat_profit = cat_sell - cat_cost
            c.setFillColor(green if cat_profit >= 0 else red)
            c.drawRightString(cols['margin'], y, f'£{cat_profit:,.2f}')
            y -= 16

            for m in cat_materials:
                y, page_num = check_space(c, y, 13, page_num)

                unit_cost = float(m.unit_cost or 0)
                unit_sell = float(m.unit_sell or 0)
                qty = int(m.quantity or 0)
                total_cost = float(m.total_cost or 0)
                total_sell = float(m.total_sell or 0)
                line_profit = total_sell - total_cost
                margin_pct = ((line_profit / total_cost) * 100) if total_cost > 0 else 0

                grand_cost += total_cost
                grand_sell += total_sell

                c.setFont('Helvetica', 7.5)
                c.setFillColor(dark)

                # Part number
                c.drawString(cols['part'], y, (m.part_number or '-')[:15])
                # Description (truncated)
                c.drawString(cols['desc'], y, (m.description or '-')[:45])
                # Qty
                c.drawString(cols['qty'], y, str(qty))
                # Unit cost
                c.drawRightString(cols['cost_t'] - 4, y, f'£{unit_cost:,.2f}')
                # Total cost
                c.drawRightString(cols['sell_u'] - 4, y, f'£{total_cost:,.2f}')
                # Unit sell
                c.drawRightString(cols['sell_t'] - 4, y, f'£{unit_sell:,.2f}')
                # Total sell
                c.drawRightString(cols['profit'] - 4, y, f'£{total_sell:,.2f}')
                # Profit + margin
                c.setFillColor(green if line_profit >= 0 else red)
                c.setFont('Helvetica-Bold', 7.5)
                c.drawRightString(cols['margin'] - 30, y, f'£{line_profit:,.2f}')
                c.setFont('Helvetica', 7)
                c.drawRightString(cols['margin'], y, f'{margin_pct:.1f}%')

                # Row line
                y -= 3
                c.setStrokeColor(line_col)
                c.setLineWidth(0.3)
                c.line(margin_left, y, margin_right, y)
                y -= 10

        # ── GRAND TOTALS ─────────────────────────────────────────
        y -= 5
        y, page_num = check_space(c, y, 60, page_num)

        grand_profit = grand_sell - grand_cost
        overall_margin = ((grand_profit / grand_cost) * 100) if grand_cost > 0 else 0

        # Totals box
        box_h = 50
        c.setFillColor(HexColor('#F9FAFB'))
        c.setStrokeColor(dark)
        c.setLineWidth(1.5)
        c.roundRect(margin_left, y - box_h + 10, content_width, box_h, 5, fill=1, stroke=1)

        ty = y
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(dark)
        c.drawString(margin_left + 10, ty, 'TOTALS')

        c.setFont('Helvetica', 9)
        c.drawString(margin_left + content_width * 0.30, ty, 'Total Cost:')
        c.setFillColor(red)
        c.drawRightString(margin_left + content_width * 0.48, ty, f'£{grand_cost:,.2f}')

        c.setFillColor(dark)
        c.drawString(margin_left + content_width * 0.50, ty, 'Total Sell:')
        c.setFillColor(dark)
        c.drawRightString(margin_left + content_width * 0.68, ty, f'£{grand_sell:,.2f}')

        ty -= 16
        c.setFont('Helvetica-Bold', 11)
        c.setFillColor(green if grand_profit >= 0 else red)
        c.drawString(margin_left + 10, ty, f'MATERIALS PROFIT: £{grand_profit:,.2f}  ({overall_margin:.1f}%)')

        # Labour + contingency + grand total
        labour_cost = float(project.total_labour_cost or 0)
        contingency = float(project.contingency_amount or 0)
        project_total = float(project.grand_total or 0)

        c.setFont('Helvetica', 9)
        c.setFillColor(dark)
        c.drawString(margin_left + content_width * 0.50, ty, f'Labour: £{labour_cost:,.2f}')
        c.drawString(margin_left + content_width * 0.72, ty, f'Contingency: £{contingency:,.2f}')

        ty -= 14
        total_profit = project_total - grand_cost - labour_cost
        c.setFont('Helvetica-Bold', 10)
        c.setFillColor(dark)
        c.drawString(margin_left + content_width * 0.50, ty, f'Project Total: £{project_total:,.2f}')
        c.setFillColor(green if total_profit >= 0 else red)
        c.drawString(margin_left + content_width * 0.72, ty, f'Est. Project Profit: £{total_profit:,.2f}')

        # Footer
        draw_footer(c, page_num)
        c.save()

        current_app.logger.info(f"Materials list PDF generated: {filepath}")

        from flask import send_file
        return send_file(
            filepath,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Materials list PDF error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to generate materials list. Please try again.'}), 500        


# =============================================================================
# TAKEOFF - DRAWING RENDERING
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/render')
@login_required
def render_document(project_id, doc_id):
    """Render a project document (PDF page) as an image for the takeoff canvas.
    
    For PDFs: converts to PNG using PyMuPDF
    For images: serves directly
    """
    from app.models.project import Project, ProjectDocument

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    page = request.args.get('page', 1, type=int)

    # Check if we already have a rendered version
    render_dir = os.path.join(current_app.root_path, 'uploads', 'projects', str(project_id), 'renders')
    os.makedirs(render_dir, exist_ok=True)
    render_path = os.path.join(render_dir, f'doc_{doc_id}_page_{page}.png')

    # Always re-render if file is empty or missing
    if not os.path.exists(render_path) or os.path.getsize(render_path) == 0:
        mime = (document.mime_type or '').lower()
        current_app.logger.info(f"Rendering doc {doc_id}, file_path={document.file_path}, exists={os.path.exists(document.file_path)}")

        if 'pdf' in mime or document.original_filename.lower().endswith('.pdf'):
            try:
                import fitz  # PyMuPDF
                if not os.path.exists(document.file_path):
                    current_app.logger.error(f"PDF file not found: {document.file_path}")
                    return jsonify({'success': False, 'error': 'PDF file not found. Please re-upload the document.'}), 404
                doc = fitz.open(document.file_path)
                if page <= len(doc):
                    pg = doc[page - 1]
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for quality
                    pix = pg.get_pixmap(matrix=mat)
                    pix.save(render_path)
                    current_app.logger.info(f"PNG saved to {render_path}, size={os.path.getsize(render_path)}")
                else:
                    return jsonify({'success': False, 'error': 'Page not found'}), 404
            except ImportError:
                return jsonify({'success': False, 'error': 'PyMuPDF not installed. Run: pip install PyMuPDF'}), 500
            except Exception as e:
                current_app.logger.error(f"PDF render error for {document.file_path}: {e}"); return jsonify({'success': False, 'error': 'Failed to render PDF. The file may be corrupted.'}), 500

        elif any(ext in mime for ext in ['png', 'jpeg', 'jpg']):
            # Already an image, just copy
            import shutil
            shutil.copy2(document.file_path, render_path)
        else:
            return jsonify({'success': False, 'error': f'Unsupported file type: {mime}'}), 400

    # Check file before sending
    if os.path.exists(render_path):
        file_size = os.path.getsize(render_path)
        current_app.logger.info(f"Sending render_path={render_path}, size={file_size}")
        if file_size == 0:
            return jsonify({'success': False, 'error': 'Rendered file is empty'}), 500
        # Use Response with file data directly
        with open(render_path, 'rb') as f:
            image_data = f.read()
        from flask import Response
        return Response(image_data, mimetype='image/png', headers={
            'Content-Length': str(len(image_data)),
            'Cache-Control': 'public, max-age=3600'
        })
    else:
        current_app.logger.error(f"Render path not found: {render_path}")
        return jsonify({'success': False, 'error': 'Rendered file not found'}), 404


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/info')
@login_required
def document_info(project_id, doc_id):
    """Get document info including page count for PDFs"""
    from app.models.project import Project, ProjectDocument

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    page_count = 1
    if document.original_filename.lower().endswith('.pdf'):
        try:
            import fitz
            doc = fitz.open(document.file_path)
            page_count = len(doc)
        except Exception:
            pass

    return jsonify({
        'success': True,
        'document': document.to_dict(),
        'page_count': page_count,
    })


# =============================================================================
# TAKEOFF - SCALE CALIBRATION
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/scale', methods=['POST'])
@login_required
def set_drawing_scale(project_id, doc_id):
    """Set the drawing scale from a known measurement."""
    from app.models.project import Project, ProjectDocument
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    data = request.get_json()
    pixel_distance = data.get('pixel_distance')
    real_distance = data.get('real_distance')
    scale_label = data.get('scale_label')

    if not pixel_distance or not real_distance or real_distance <= 0:
        return jsonify({'success': False, 'error': 'Invalid measurements'}), 400

    px_per_metre = pixel_distance / real_distance
    document.scale = json.dumps({
        'px_per_metre': round(px_per_metre, 2),
        'label': scale_label,
        'calibration_px': pixel_distance,
        'calibration_m': real_distance,
    })

    db.session.commit()

    return jsonify({
        'success': True,
        'scale': {
            'px_per_metre': round(px_per_metre, 2),
            'label': scale_label,
        }
    })


def _get_scale(document):
    if document.scale:
        try:
            # Handle both JSON string and raw float
            if isinstance(document.scale, (int, float)):
                return float(document.scale)
            data = json.loads(document.scale)
            if isinstance(data, (int, float)):
                return float(data)
            return data.get('px_per_metre', 50)
        except (json.JSONDecodeError, TypeError, AttributeError, ValueError):
            try:
                return float(document.scale)
            except (TypeError, ValueError):
                pass
    return 50

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/auto-scale', methods=['POST'])
@login_required
def auto_scale(project_id, doc_id):
    """Calculate scale automatically from drawing notation like '1:35 @ A1'.
    
    Request body:
        {
            "scale_ratio": 35,       # The '35' from '1:35'
            "paper_size": "A1",      # Paper size
            "orientation": "landscape"  # or "portrait" (default: landscape for drawings)
        }
    
    Returns:
        { "success": true, "px_per_metre": 168.5, "description": "1:35 @ A1 landscape" }
    """
    from app.models.project import Project, ProjectDocument
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    data = request.get_json()
    scale_ratio = data.get('scale_ratio')  # e.g. 35 for 1:35
    paper_size = data.get('paper_size', 'A1').upper()
    orientation = data.get('orientation', 'landscape')

    if not scale_ratio or scale_ratio <= 0:
        return jsonify({'success': False, 'error': 'Invalid scale ratio'}), 400

    if paper_size not in PAPER_SIZES_MM:
        return jsonify({'success': False, 'error': f'Unknown paper size: {paper_size}'}), 400

    # Get paper dimensions
    paper_w_mm, paper_h_mm = PAPER_SIZES_MM[paper_size]
    if orientation == 'landscape':
        paper_long_mm = max(paper_w_mm, paper_h_mm)
    else:
        paper_long_mm = min(paper_w_mm, paper_h_mm)

    # Real-world distance that the paper width represents
    real_width_mm = paper_long_mm * scale_ratio
    real_width_m = real_width_mm / 1000.0

    # Get rendered image dimensions
    render_dir = os.path.join(current_app.root_path, 'uploads', 'projects', str(project_id), 'renders')
    render_path = os.path.join(render_dir, f'doc_{doc_id}_page_1.png')

    if not os.path.exists(render_path):
        return jsonify({'success': False, 'error': 'Drawing not rendered yet'}), 400

    import cv2
    img = cv2.imread(render_path)
    if img is None:
        return jsonify({'success': False, 'error': 'Could not read rendered image'}), 500

    img_width_px = img.shape[1]  # Width in pixels
    img_height_px = img.shape[0]

    # Use the longer dimension (usually landscape drawings)
    img_long_px = max(img_width_px, img_height_px)

    # Calculate pixels per metre
    px_per_metre = img_long_px / real_width_m

    # Save to document
    document.scale = px_per_metre
    db.session.commit()

    description = f"1:{scale_ratio} @ {paper_size} {orientation}"

    current_app.logger.info(
        f"Auto-scale: {description} → paper={paper_long_mm}mm, "
        f"real_width={real_width_m:.2f}m, img={img_long_px}px, "
        f"px_per_metre={px_per_metre:.2f}"
    )

    return jsonify({
        'success': True,
        'px_per_metre': round(px_per_metre, 2),
        'real_width_m': round(real_width_m, 2),
        'description': description,
    })    


# =============================================================================
# TAKEOFF - SYMBOL TEMPLATES
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/symbol-templates', methods=['GET'])
@login_required
def get_symbol_templates(project_id, doc_id):
    """Get all symbol templates for a document"""
    from app.models.takeoff import TakeoffSymbolTemplate

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    templates = TakeoffSymbolTemplate.query.filter_by(
        project_id=project_id, document_id=doc_id
    ).all()

    return jsonify({
        'success': True,
        'templates': [t.to_dict() for t in templates],
    })


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/symbol-templates', methods=['POST'])
@login_required
def create_symbol_template(project_id, doc_id):
    """Create a symbol template from a user-drawn bounding box on the key area."""
    from app.models.project import Project, ProjectDocument
    from app.models.takeoff import TakeoffSymbolTemplate
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    label = str(data.get('label', 'Unknown Symbol'))[:100].strip()
    crop_x = data.get('crop_x')
    crop_y = data.get('crop_y')
    crop_w = data.get('crop_w')
    crop_h = data.get('crop_h')
    crop_image_b64 = data.get('crop_image')

    if not all([crop_x is not None, crop_y is not None, crop_w, crop_h]):
        return jsonify({'success': False, 'error': 'Crop coordinates required'}), 400

    # Validate crop coordinates are numbers
    try:
        crop_x = int(crop_x)
        crop_y = int(crop_y)
        crop_w = int(crop_w)
        crop_h = int(crop_h)
        if crop_w <= 0 or crop_h <= 0 or crop_w > 2000 or crop_h > 2000:
            return jsonify({'success': False, 'error': 'Invalid crop dimensions'}), 400
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Crop coordinates must be numbers'}), 400
    
    # Validate base64 image size (max 5MB decoded)
    if crop_image_b64:
        raw_b64 = crop_image_b64.split(',')[1] if ',' in crop_image_b64 else crop_image_b64
        if len(raw_b64) > 5 * 1024 * 1024 * 1.37:  # ~5MB after base64 encoding overhead
            return jsonify({'success': False, 'error': 'Crop image too large'}), 400
    
    # Limit templates per document (max 50)
    template_count = TakeoffSymbolTemplate.query.filter_by(project_id=project_id, document_id=doc_id).count()
    if template_count >= 50:
        return jsonify({'success': False, 'error': 'Maximum 50 symbol templates per document'}), 400

    symbol_type_id = f"sym_{uuid.uuid4().hex[:8]}"

    crop_image_path = None
    if crop_image_b64:
        crop_dir = os.path.join(current_app.root_path, 'uploads', 'projects', str(project_id), 'symbols')
        os.makedirs(crop_dir, exist_ok=True)
        crop_image_path = os.path.join(crop_dir, f'{symbol_type_id}.png')

        try:
            if ',' in crop_image_b64:
                crop_image_b64 = crop_image_b64.split(',')[1]
            img_data = base64.b64decode(crop_image_b64)
            with open(crop_image_path, 'wb') as f:
                f.write(img_data)
        except Exception as e:
            current_app.logger.warning(f"Could not save crop image: {e}")
            crop_image_path = None

    template = TakeoffSymbolTemplate(
        project_id=project.id,
        document_id=document.id,
        symbol_type_id=symbol_type_id,
        label=label,
        crop_x=crop_x,
        crop_y=crop_y,
        crop_w=crop_w,
        crop_h=crop_h,
        crop_image_path=crop_image_path,
        color=data.get('color', '#3b82f6'),
        icon=data.get('icon'),
    )

    db.session.add(template)
    db.session.commit()

    return jsonify({
        'success': True,
        'template': template.to_dict(),
    })


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/symbol-templates/<int:template_id>', methods=['DELETE'])
@login_required
def delete_symbol_template(project_id, doc_id, template_id):
    """Delete a symbol template and all its detections"""
    from app.models.takeoff import TakeoffSymbolTemplate, TakeoffSymbolDetection
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    template = TakeoffSymbolTemplate.query.filter_by(
        id=template_id, project_id=project_id, document_id=doc_id
    ).first()
    if not template:
        return jsonify({'success': False, 'error': 'Template not found'}), 404

    TakeoffSymbolDetection.query.filter_by(
        project_id=project_id, symbol_type_id=template.symbol_type_id
    ).delete()

    db.session.delete(template)
    db.session.commit()

    return jsonify({'success': True})


# =============================================================================
# TAKEOFF - SYMBOL DETECTION
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/detect-symbols', methods=['POST'])
@login_required
@rate_limit(5, 60)
def detect_symbols(project_id, doc_id):
    """Run symbol detection on a drawing for a given symbol template."""
    from app.models.project import Project, ProjectDocument
    from app.models.takeoff import TakeoffSymbolTemplate, TakeoffSymbolDetection, TakeoffRoom
    from app.extensions import db
    from app.services.symbol_detector import SymbolDetector

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    data = request.get_json()
    template_id = data.get('template_id')
    exclude_area = data.get('exclude_area')
    confidence_threshold = data.get('confidence_threshold', 0.7)
    page = data.get('page', 1)

    template = TakeoffSymbolTemplate.query.filter_by(
        id=template_id, project_id=project_id
    ).first()
    if not template:
        return jsonify({'success': False, 'error': 'Symbol template not found'}), 404

    render_dir = os.path.join(current_app.root_path, 'uploads', 'projects', str(project_id), 'renders')
    render_path = os.path.join(render_dir, f'doc_{doc_id}_page_{page}.png')

    if not os.path.exists(render_path):
        return jsonify({'success': False, 'error': 'Drawing not rendered yet. Open the takeoff view first.'}), 400

    try:
        detector = SymbolDetector()

        # Step 1: OpenCV finds ALL candidates at any rotation (accurate positions)
        detections = detector.detect(
            drawing_path=render_path,
            template_path=template.crop_image_path,
            crop_rect={'x': template.crop_x, 'y': template.crop_y,
                       'w': template.crop_w, 'h': template.crop_h},
            exclude_area=exclude_area,
            confidence_threshold=confidence_threshold,
        )

        current_app.logger.info(f"OpenCV found {len(detections)} candidates for '{template.label}'")

        api_key = current_app.config.get('ANTHROPIC_API_KEY') or os.environ.get('ANTHROPIC_API_KEY')

        # Step 2: If OpenCV found candidates, verify with Claude Vision
        if detections and len(detections) > 0 and api_key:
            try:
                import anthropic
                import cv2

                client = anthropic.Anthropic(api_key=api_key)

                template_b64 = None
                if template.crop_image_path and os.path.exists(template.crop_image_path):
                    with open(template.crop_image_path, 'rb') as f:
                        template_b64 = base64.b64encode(f.read()).decode()

                drawing = cv2.imread(render_path)

                candidate_crops = []
                for i, det in enumerate(detections):
                    pad = 10
                    bx = max(0, det.get('box_x', det['x'] - det['w']//2) - pad)
                    by = max(0, det.get('box_y', det['y'] - det['h']//2) - pad)
                    bw = min(drawing.shape[1] - bx, det['w'] + pad * 2)
                    bh = min(drawing.shape[0] - by, det['h'] + pad * 2)
                    crop = drawing[by:by+bh, bx:bx+bw]
                    if crop.size == 0:
                        continue
                    _, buf = cv2.imencode('.png', crop)
                    crop_b64 = base64.b64encode(buf).decode()
                    candidate_crops.append({'index': i, 'b64': crop_b64})

                if candidate_crops and template_b64:
                    content = []
                    content.append({
                        "type": "text",
                        "text": (f"I have a reference electrical symbol from a drawing key. "
                                 f"Below it are {len(candidate_crops)} candidate regions found on the drawing. "
                                 f"Tell me which candidates VISUALLY match the reference symbol — "
                                 f"same shape AND same letter/text inside. "
                                 f"Candidates may be ROTATED at any angle. "
                                 f"The text/annotation near symbols may be oriented differently from the symbol itself — ignore text orientation. "
                                 f"Similar symbols with DIFFERENT letters (e.g. S vs M vs H inside circles) are NOT matches.\n\n"
                                 f"Reply with ONLY a JSON array of matching candidate indices, e.g. [0, 2, 5]\n"
                                 f"If none match: []")
                    })
                    content.append({"type": "text", "text": "Reference symbol:"})
                    content.append({
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/png", "data": template_b64}
                    })

                    for cc in candidate_crops:
                        content.append({"type": "text", "text": f"Candidate #{cc['index']}:"})
                        content.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": "image/png", "data": cc['b64']}
                        })

                    response = client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=500,
                        messages=[{"role": "user", "content": content}],
                    )

                    response_text = response.content[0].text.strip()
                    current_app.logger.info(f"Claude Vision verification: {response_text}")

                    import re
                    json_match = re.search(r'\[.*?\]', response_text)
                    if json_match:
                        matching_indices = json.loads(json_match.group())
                        original_count = len(detections)
                        detections = [detections[i] for i in matching_indices if i < len(detections)]
                        current_app.logger.info(
                            f"Claude Vision: {original_count} candidates → {len(detections)} verified"
                        )

            except Exception as vision_err:
                current_app.logger.warning(
                    f"Claude Vision verification failed (keeping OpenCV results): {vision_err}"
                )

        # Step 3: If OpenCV found few results, use Claude Vision GRID search
        # This handles complex symbols where template matching fails
        # (symbols with text that rotates separately, complex multi-part symbols)
        if len(detections) < 2 and api_key and template.crop_image_path and os.path.exists(template.crop_image_path):
            try:
                import anthropic
                import cv2

                client = anthropic.Anthropic(api_key=api_key)

                drawing = cv2.imread(render_path)
                img_h, img_w = drawing.shape[:2]

                with open(template.crop_image_path, 'rb') as f:
                    template_b64 = base64.b64encode(f.read()).decode()

                # Split drawing into overlapping tiles
                # Each tile is ~800x800 px with 200px overlap
                tile_size = 800
                overlap_px = 200
                step = tile_size - overlap_px
                tiles = []

                for ty in range(0, img_h, step):
                    for tx in range(0, img_w, step):
                        tx2 = min(tx + tile_size, img_w)
                        ty2 = min(ty + tile_size, img_h)
                        if (tx2 - tx) < 200 or (ty2 - ty) < 200:
                            continue
                        # Skip tiles entirely within key area
                        if exclude_area:
                            ka = exclude_area
                            if (tx >= ka['x'] and tx2 <= ka['x'] + ka['w'] and
                                ty >= ka['y'] and ty2 <= ka['y'] + ka['h']):
                                continue
                        tiles.append({'x': tx, 'y': ty, 'x2': tx2, 'y2': ty2})

                current_app.logger.info(f"Grid search: {len(tiles)} tiles for '{template.label}'")

                grid_detections = []
                # Process tiles in batches (up to 4 tiles per API call)
                batch_size = 4
                for batch_start in range(0, len(tiles), batch_size):
                    batch_tiles = tiles[batch_start:batch_start + batch_size]

                    content = []
                    content.append({
                        "type": "text",
                        "text": (f"Find ALL instances of the reference symbol in the following {len(batch_tiles)} tile(s) "
                                 f"from an electrical drawing. The symbol may be ROTATED at any angle on the drawing. "
                                 f"Text annotations near symbols may face a different direction — this is normal, "
                                 f"match by the GRAPHICAL SHAPE not text orientation.\n\n"
                                 f"For each tile, report found symbols with x,y coordinates RELATIVE to that tile's top-left corner (in pixels).\n"
                                 f"Reply with ONLY JSON: {{\"tile_0\": [{{\"x\":..., \"y\":...}}], \"tile_1\": [...], ...}}\n"
                                 f"Use empty arrays for tiles with no matches. No other text.")
                    })
                    content.append({"type": "text", "text": "Reference symbol:"})
                    content.append({
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/png", "data": template_b64}
                    })

                    for idx, tile in enumerate(batch_tiles):
                        crop = drawing[tile['y']:tile['y2'], tile['x']:tile['x2']]
                        _, buf = cv2.imencode('.png', crop)
                        tile_b64 = base64.b64encode(buf).decode()
                        tw = tile['x2'] - tile['x']
                        th = tile['y2'] - tile['y']
                        content.append({
                            "type": "text",
                            "text": f"Tile {idx} ({tw}x{th} pixels):"
                        })
                        content.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": "image/png", "data": tile_b64}
                        })

                    response = client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=1000,
                        messages=[{"role": "user", "content": content}],
                    )

                    response_text = response.content[0].text.strip()
                    current_app.logger.info(f"Grid batch {batch_start//batch_size}: {response_text[:300]}")

                    # Parse response
                    import re
                    json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                    if json_match:
                        try:
                            tile_results = json.loads(json_match.group())
                            for idx, tile in enumerate(batch_tiles):
                                key = f"tile_{idx}"
                                if key in tile_results and tile_results[key]:
                                    for det_item in tile_results[key]:
                                        # Convert tile-relative coords to full image coords
                                        abs_x = tile['x'] + int(det_item['x'])
                                        abs_y = tile['y'] + int(det_item['y'])

                                        # Skip if in key area
                                        if exclude_area:
                                            ka = exclude_area
                                            if (ka['x'] <= abs_x <= ka['x'] + ka['w'] and
                                                ka['y'] <= abs_y <= ka['y'] + ka['h']):
                                                continue

                                        grid_detections.append({
                                            'x': abs_x, 'y': abs_y,
                                            'w': template.crop_w or 30,
                                            'h': template.crop_h or 30,
                                            'confidence': 0.85,
                                            'box_x': abs_x - (template.crop_w or 30) // 2,
                                            'box_y': abs_y - (template.crop_h or 30) // 2,
                                            'source': 'claude_grid',
                                        })
                        except json.JSONDecodeError:
                            current_app.logger.warning(f"Failed to parse grid batch response")

                # Deduplicate grid detections (merge within 50px of each other or existing)
                merged = []
                for gd in grid_detections:
                    is_dup = False
                    for existing in merged + detections:
                        dist = ((gd['x'] - existing['x'])**2 + (gd['y'] - existing['y'])**2) ** 0.5
                        if dist < 50:
                            is_dup = True
                            break
                    if not is_dup:
                        merged.append(gd)

                if merged:
                    current_app.logger.info(f"Grid search found {len(merged)} additional detections for '{template.label}'")
                    detections.extend(merged)
                else:
                    current_app.logger.info(f"Grid search found no additional detections")

            except Exception as grid_err:
                current_app.logger.warning(f"Grid search failed: {grid_err}", exc_info=True)

    except Exception as e:
        current_app.logger.error(f"Symbol detection error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Symbol detection failed. Please try again.'}), 500

    # Delete old detections for this symbol type only
    TakeoffSymbolDetection.query.filter_by(
        project_id=project_id,
        document_id=doc_id,
        symbol_type_id=template.symbol_type_id,
    ).delete()
    db.session.flush()

    # Get ALL existing detections on this document (from other symbol types)
    # to prevent overlapping boxes from different scans
    existing_detections = TakeoffSymbolDetection.query.filter_by(
        project_id=project_id,
        document_id=doc_id,
        rejected=False,
    ).all()
    existing_coords = [(d.x, d.y) for d in existing_detections]

    MIN_OVERLAP_DISTANCE = 30  # pixels - skip if another box is within this radius

    rooms = TakeoffRoom.query.filter_by(project_id=project_id, document_id=doc_id).all()

    new_detections = []
    skipped = 0
    for det in detections:
        # Check overlap with ANY existing detection (from other symbol types)
        too_close = False
        for ex, ey in existing_coords:
            dist = ((det['x'] - ex) ** 2 + (det['y'] - ey) ** 2) ** 0.5
            if dist < MIN_OVERLAP_DISTANCE:
                too_close = True
                break
        if too_close:
            skipped += 1
            continue

        room_id = None
        for room in rooms:
            points = room.get_boundary_points()
            if points and _point_in_polygon(det['x'], det['y'], points):
                room_id = room.id
                break
            if room.bbox_x and room.bbox_y:
                if (room.bbox_x <= det['x'] <= room.bbox_x + room.bbox_w and
                    room.bbox_y <= det['y'] <= room.bbox_y + room.bbox_h):
                    room_id = room.id
                    break

        detection = TakeoffSymbolDetection(
            project_id=project_id,
            document_id=doc_id,
            room_id=room_id,
            symbol_type_id=template.symbol_type_id,
            symbol_label=template.label,
            x=det['x'],
            y=det['y'],
            confidence=det.get('confidence', 0.8),
            confirmed=False,
            source='opencv',
            part_number=template.default_part_number,
            product_description=template.default_product_description,
        )
        db.session.add(detection)
        new_detections.append(detection)

    template.total_found = len(new_detections)
    if skipped:
        current_app.logger.info(f"Skipped {skipped} overlapping detections for {template.label}")

    db.session.commit()

    return jsonify({
        'success': True,
        'count': len(new_detections),
        'skipped': skipped,
        'detections': [d.to_dict() for d in new_detections],
    })


def _point_in_polygon(x, y, polygon):
    """Ray casting algorithm to check if point is inside polygon"""
    n = len(polygon)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = polygon[i]['x'], polygon[i]['y']
        xj, yj = polygon[j]['x'], polygon[j]['y']
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/detect-symbols-ai', methods=['POST'])
@login_required
@rate_limit(3, 60)
def detect_symbols_ai(project_id, doc_id):
    """
    AI-powered symbol detection using Claude Vision API.
    Handles colours, text inside symbols, gang counts, dimmers.
    """
    from app.models.project import Project, ProjectDocument
    from app.models.takeoff import TakeoffSymbolTemplate, TakeoffSymbolDetection, TakeoffRoom
    from app.extensions import db
    
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404
    
    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404
    
    data = request.get_json() or {}
    room_id = data.get('room_id')  # Optional: focus on specific room
    
    # Get all symbol templates for this project
    templates = TakeoffSymbolTemplate.query.filter_by(
        project_id=project_id, document_id=doc_id
    ).all()
    
    if not templates:
        return jsonify({'success': False, 'error': 'No symbol templates defined. Create templates first.'}), 400
    
    # Build template data with descriptions for Claude
    template_data = []
    for t in templates:
        template_info = {
            'id': t.id,
            'name': t.label or f'Symbol {t.id}',
            'category': t.symbol_type_id or 'other',
            'description': '',
            'distinguishing_features': []
        }
        
        # Add distinguishing features if available
        features = []
        if hasattr(t, 'colour') and t.colour:
            features.append(f"colour: {t.colour}")
        if hasattr(t, 'expected_text') and t.expected_text:
            features.append(f"has letter '{t.expected_text}' inside")
        if hasattr(t, 'gang_count') and t.gang_count:
            features.append(f"{t.gang_count}-gang (has {t.gang_count} tick marks)")
        if hasattr(t, 'is_dimmer') and t.is_dimmer:
            features.append("is a dimmer (has 'D' subscript)")
        
        template_info['distinguishing_features'] = ', '.join(features) if features else None
        template_data.append(template_info)
    
    try:
        from app.services.symbol_detector_ai import detect_symbols_with_ai
        
        detections = detect_symbols_with_ai(
            pdf_path=document.file_path,
            page_num=0,
            symbol_templates=template_data,
            zoom=2.0
        )
        
        # Get rooms for assignment
        rooms = TakeoffRoom.query.filter_by(project_id=project_id, document_id=doc_id).all()
        
        # Assign detections to rooms based on position
        for detection in detections:
            detection['room_id'] = None
            center_x = detection['x'] + detection['width'] / 2
            center_y = detection['y'] + detection['height'] / 2
            
            for room in rooms:
                points = room.get_boundary_points() if hasattr(room, 'get_boundary_points') else None
                if points and _point_in_polygon(center_x, center_y, points):
                    detection['room_id'] = room.id
                    break
        
        # Filter by room if specified
        if room_id:
            detections = [d for d in detections if d.get('room_id') == room_id]
        
        # Save to database
        # Get existing detections to prevent overlapping boxes
        existing_detections = TakeoffSymbolDetection.query.filter_by(
            project_id=project_id,
            document_id=doc_id,
            rejected=False,
        ).all()
        existing_coords = [(ed.x, ed.y) for ed in existing_detections]
        MIN_OVERLAP_DISTANCE = 30

        saved_detections = []
        for d in detections:
            # Skip if overlapping with existing detection
            too_close = False
            for ex, ey in existing_coords:
                dist = ((d['x'] - ex) ** 2 + (d['y'] - ey) ** 2) ** 0.5
                if dist < MIN_OVERLAP_DISTANCE:
                    too_close = True
                    break
            if too_close:
                continue

            # Find matching template by name
            template_id = d.get('template_id')
            template = None
            if template_id:
                template = TakeoffSymbolTemplate.query.get(template_id)
            if not template:
                # Try to match by name
                template_name = d.get('template_name', '')
                for t in templates:
                    if t.label and template_name and t.label.lower() in template_name.lower():
                        template = t
                        break
            
            detection = TakeoffSymbolDetection(
                project_id=project_id,
                document_id=doc_id,
                room_id=d.get('room_id'),
                symbol_type_id=template.symbol_type_id if template else None,
                symbol_label=d.get('template_name') or (template.label if template else 'Unknown'),
                x=d['x'],
                y=d['y'],
                confidence=d.get('confidence', 0.8),
                confirmed=False,
                source='ai_vision',
                part_number=template.default_part_number if template else None,
                product_description=template.default_product_description if template else None,
            )
            
            db.session.add(detection)
            db.session.flush()
            
            saved_detections.append({
                'id': detection.id,
                'x': detection.x,
                'y': detection.y,
                'width': d.get('width', 40),
                'height': d.get('height', 40),
                'confidence': detection.confidence,
                'template_id': template.id if template else None,
                'template_name': detection.symbol_label,
                'room_id': detection.room_id,
                'colour': d.get('colour'),
                'text_inside': d.get('text_inside'),
                'gang_count': d.get('gang_count'),
                'is_dimmer': d.get('is_dimmer'),
                'location_description': d.get('location_description'),
                'notes': d.get('notes')
            })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'detections': saved_detections,
            'total_found': len(saved_detections),
            'method': 'ai_vision'
        })
        
    except Exception as e:
        current_app.logger.error(f"AI symbol detection failed: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'error': 'AI symbol detection failed. Please try again.'}), 500
# =============================================================================
# TAKEOFF - MANUAL SYMBOL PLACEMENT
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/detections', methods=['POST'])
@login_required
def add_manual_detection(project_id, doc_id):
    """Manually place a symbol detection on the drawing"""
    from app.models.takeoff import TakeoffSymbolDetection
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()

    detection = TakeoffSymbolDetection(
        project_id=project_id,
        document_id=doc_id,
        room_id=data.get('room_id'),
        symbol_type_id=data.get('symbol_type_id'),
        symbol_label=data.get('symbol_label'),
        x=data['x'],
        y=data['y'],
        confidence=1.0,
        confirmed=True,
        source='manual',
    )

    db.session.add(detection)
    db.session.commit()

    return jsonify({'success': True, 'detection': detection.to_dict()})


@bp.route('/api/projects/<int:project_id>/detections/<int:detection_id>', methods=['PUT'])
@login_required
def update_detection(project_id, detection_id):
    """Update a detection (confirm, reject, change room, link product)"""
    from app.models.takeoff import TakeoffSymbolDetection
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    detection = TakeoffSymbolDetection.query.filter_by(
        id=detection_id, project_id=project_id
    ).first()
    if not detection:
        return jsonify({'success': False, 'error': 'Detection not found'}), 404

    data = request.get_json()
    for field in ['confirmed', 'rejected', 'room_id', 'part_number', 'product_description', 'material_id', 'x', 'y']:
        if field in data:
            setattr(detection, field, data[field])

    db.session.commit()
    return jsonify({'success': True, 'detection': detection.to_dict()})


@bp.route('/api/projects/<int:project_id>/detections/<int:detection_id>', methods=['DELETE'])
@login_required
def delete_detection(project_id, detection_id):
    """Delete a single detection"""
    from app.models.takeoff import TakeoffSymbolDetection
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    detection = TakeoffSymbolDetection.query.filter_by(
        id=detection_id, project_id=project_id
    ).first()
    if not detection:
        return jsonify({'success': False, 'error': 'Detection not found'}), 404

    db.session.delete(detection)
    db.session.commit()
    return jsonify({'success': True})


# =============================================================================
# TAKEOFF - ROOMS
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/rooms', methods=['GET'])
@login_required
def get_rooms(project_id, doc_id):
    """Get all rooms for a document"""
    from app.models.takeoff import TakeoffRoom

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    rooms = TakeoffRoom.query.filter_by(
        project_id=project_id, document_id=doc_id
    ).order_by(TakeoffRoom.sort_order).all()

    return jsonify({
        'success': True,
        'rooms': [r.to_dict() for r in rooms],
    })


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/rooms', methods=['POST'])
@login_required
def create_room(project_id, doc_id):
    """Create a room zone on the drawing"""
    from app.models.takeoff import TakeoffRoom
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()

    room = TakeoffRoom(
        project_id=project_id,
        document_id=doc_id,
        name=data.get('name', 'New Room'),
        floor_level=data.get('floor_level'),
        room_type=data.get('room_type'),
        color=data.get('color', '#6366f1'),
    )

    points = data.get('boundary_points', [])
    room.set_boundary_points(points)

    if points:
        xs = [p['x'] for p in points]
        ys = [p['y'] for p in points]
        room.bbox_x = min(xs)
        room.bbox_y = min(ys)
        room.bbox_w = max(xs) - min(xs)
        room.bbox_h = max(ys) - min(ys)

    db.session.add(room)
    db.session.commit()

    _reassign_detections_to_rooms(project_id, doc_id)

    return jsonify({'success': True, 'room': room.to_dict()})


@bp.route('/api/projects/<int:project_id>/rooms/<int:room_id>', methods=['PUT'])
@login_required
def update_room(project_id, room_id):
    """Update a room"""
    from app.models.takeoff import TakeoffRoom
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    room = TakeoffRoom.query.filter_by(id=room_id, project_id=project_id).first()
    if not room:
        return jsonify({'success': False, 'error': 'Room not found'}), 404

    data = request.get_json()
    for field in ['name', 'floor_level', 'room_type', 'color', 'sort_order']:
        if field in data:
            setattr(room, field, data[field])

    if 'boundary_points' in data:
        room.set_boundary_points(data['boundary_points'])
        points = data['boundary_points']
        if points:
            xs = [p['x'] for p in points]
            ys = [p['y'] for p in points]
            room.bbox_x = min(xs)
            room.bbox_y = min(ys)
            room.bbox_w = max(xs) - min(xs)
            room.bbox_h = max(ys) - min(ys)

    db.session.commit()
    _reassign_detections_to_rooms(project_id, room.document_id)

    return jsonify({'success': True, 'room': room.to_dict()})


@bp.route('/api/projects/<int:project_id>/rooms/<int:room_id>', methods=['DELETE'])
@login_required
def delete_room(project_id, room_id):
    """Delete a room and unassign its detections"""
    from app.models.takeoff import TakeoffRoom, TakeoffSymbolDetection
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    room = TakeoffRoom.query.filter_by(id=room_id, project_id=project_id).first()
    if not room:
        return jsonify({'success': False, 'error': 'Room not found'}), 404

    TakeoffSymbolDetection.query.filter_by(room_id=room.id).update({'room_id': None})

    db.session.delete(room)
    db.session.commit()

    return jsonify({'success': True})


def _reassign_detections_to_rooms(project_id, doc_id):
    """Re-assign all detections to rooms based on their position"""
    from app.models.takeoff import TakeoffRoom, TakeoffSymbolDetection
    from app.extensions import db

    rooms = TakeoffRoom.query.filter_by(project_id=project_id, document_id=doc_id).all()
    detections = TakeoffSymbolDetection.query.filter_by(project_id=project_id, document_id=doc_id).all()

    for det in detections:
        det.room_id = None
        for room in rooms:
            points = room.get_boundary_points()
            if points and _point_in_polygon(det.x, det.y, points):
                det.room_id = room.id
                break
            if room.bbox_x and room.bbox_y:
                if (room.bbox_x <= det.x <= room.bbox_x + room.bbox_w and
                    room.bbox_y <= det.y <= room.bbox_y + room.bbox_h):
                    det.room_id = room.id
                    break

    db.session.commit()


# =============================================================================
# TAKEOFF - CABLE RUNS
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/cable-runs', methods=['GET'])
@login_required
def get_cable_runs(project_id, doc_id):
    """Get all cable runs for a document"""
    from app.models.takeoff import TakeoffCableRun

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    runs = TakeoffCableRun.query.filter_by(
        project_id=project_id, document_id=doc_id
    ).all()

    return jsonify({
        'success': True,
        'cable_runs': [r.to_dict() for r in runs],
    })


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/cable-runs', methods=['POST'])
@login_required
def create_cable_run(project_id, doc_id):
    """Create a new cable run from click-to-click points"""
    from app.models.project import ProjectDocument
    from app.models.takeoff import TakeoffCableRun
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()
    points = data.get('route_points', [])

    if len(points) < 2:
        return jsonify({'success': False, 'error': 'Need at least 2 points'}), 400

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    scale = _get_scale(document) if document else 50

    run = TakeoffCableRun(
        project_id=project_id,
        document_id=doc_id,
        room_id=data.get('room_id'),
        cable_type=data.get('cable_type', 'socket'),
        cable_label=data.get('cable_label'),
        waste_percent=data.get('waste_percent', 10),
        notes=data.get('notes'),
        circuit_ref=data.get('circuit_ref'),
        part_number=data.get('part_number'),
    )
    run.set_route_points(points)
    run.calculate_length(scale)

    db.session.add(run)
    db.session.commit()

    return jsonify({'success': True, 'cable_run': run.to_dict()})


@bp.route('/api/projects/<int:project_id>/cable-runs/<int:run_id>', methods=['DELETE'])
@login_required
def delete_cable_run(project_id, run_id):
    """Delete a cable run"""
    from app.models.takeoff import TakeoffCableRun
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    run = TakeoffCableRun.query.filter_by(id=run_id, project_id=project_id).first()
    if not run:
        return jsonify({'success': False, 'error': 'Cable run not found'}), 404

    db.session.delete(run)
    db.session.commit()
    return jsonify({'success': True})


# =============================================================================
# TAKEOFF - FLOOR AREAS
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/areas', methods=['POST'])
@login_required
def create_area(project_id, doc_id):
    """Create a floor area measurement"""
    from app.models.project import ProjectDocument
    from app.models.takeoff import TakeoffArea
    from app.extensions import db

    if not _verify_project_ownership(project_id):
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()
    points = data.get('points', [])

    if len(points) < 3:
        return jsonify({'success': False, 'error': 'Need at least 3 points'}), 400

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    scale = _get_scale(document) if document else 50

    area = TakeoffArea(
        project_id=project_id,
        document_id=doc_id,
        room_id=data.get('room_id'),
        label=data.get('label', 'Area'),
    )
    area.set_points(points)
    area.calculate_area(scale)

    db.session.add(area)
    db.session.commit()

    return jsonify({'success': True, 'area': area.to_dict()})


@bp.route('/api/projects/<int:project_id>/link-ufh', methods=['POST'])
@login_required
def link_ufh_product(project_id):
    """Link an underfloor heating mat product to an area measurement"""
    from app.models.project import Project
    from app.models.project import ProjectMaterial
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()
    product = data.get('product', {})
    mat_size = data.get('mat_size_sqm', 0)
    total_area = data.get('total_area_sqm', 0)
    area_id = data.get('area_id')

    category = f"Underfloor Heating - {mat_size}m²"

    material = ProjectMaterial(
        project_id=project_id,
        category=category,
        part_number=product.get('sku'),
        description=f"{product.get('description') or product.get('name')} ({mat_size}m² mat)",
        quantity=1,
        unit='each',
        unit_cost=product.get('purchase_cost'),
        price_source='quickbooks' if product.get('id') else 'manual',
        price_verified=True if product.get('id') else False,
        qb_item_id=product.get('id'),
        qb_item_name=product.get('name'),
    )
    material.calculate_totals(markup_percent=float(project.materials_markup_percent))
    qb_sell = product.get('unit_price', 0)
    if qb_sell and qb_sell > float(material.unit_sell or 0):
        material.unit_sell = qb_sell
        material.total_sell = round(1 * qb_sell, 2)

    db.session.add(material)
    project.recalculate_totals()
    db.session.commit()

    return jsonify({
        'success': True,
        'material_id': material.id,
        'mat_size_sqm': mat_size,
        'total_area_sqm': total_area,
    })    

# =============================================================================
# TAKEOFF - ACCESSORIES
# =============================================================================

@bp.route('/api/projects/<int:project_id>/accessories', methods=['POST'])
@login_required
def add_accessory(project_id):
    """Add an accessory product to a symbol template"""
    from app.models.project import Project
    from app.models.project import ProjectMaterial
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()
    product = data.get('product', {})
    quantity = data.get('quantity', 1)
    template_id = data.get('template_id')
    symbol_type_id = data.get('symbol_type_id')
    parent_label = data.get('parent_label', '')

    category = f"{parent_label} - Accessories"

    material = ProjectMaterial(
        project_id=project_id,
        category=category,
        part_number=product.get('sku'),
        description=product.get('description') or product.get('name'),
        quantity=quantity,
        unit='each',
        unit_cost=product.get('purchase_cost'),
        price_source='quickbooks' if product.get('id') else 'manual',
        price_verified=True if product.get('id') else False,
        qb_item_id=product.get('id'),
        qb_item_name=product.get('name'),
    )
    material.calculate_totals(markup_percent=float(project.materials_markup_percent))
    qb_sell = product.get('unit_price', 0)
    if qb_sell and qb_sell > float(material.unit_sell or 0):
        material.unit_sell = qb_sell
        material.total_sell = round(quantity * qb_sell, 2)

    db.session.add(material)
    project.recalculate_totals()
    db.session.commit()

    return jsonify({
        'success': True,
        'accessory': {
            'id': material.id,
            'template_id': template_id,
            'symbol_type_id': symbol_type_id,
            'part_number': material.part_number,
            'description': material.description,
            'quantity': material.quantity,
            'unit_cost': float(material.unit_cost or 0),
            'unit_sell': float(material.unit_sell or 0),
        }
    })


@bp.route('/api/projects/<int:project_id>/accessories/<int:material_id>', methods=['DELETE'])
@login_required
def remove_accessory(project_id, material_id):
    """Remove an accessory material"""
    from app.models.project import Project
    from app.models.project import ProjectMaterial
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    material = ProjectMaterial.query.filter_by(id=material_id, project_id=project_id).first()
    if not material:
        return jsonify({'success': False, 'error': 'Material not found'}), 404

    db.session.delete(material)
    project.recalculate_totals()
    db.session.commit()

    return jsonify({'success': True})


# =============================================================================
# TAKEOFF - PRODUCT SEARCH
# =============================================================================
# PRODUCT SEARCH WITH CACHING
# =============================================================================

# In-memory product cache: { user_id: { 'products': [...], 'cached_at': time } }
_product_cache = {}
_CACHE_TTL_SECONDS = 600  # 10 minutes


@bp.route('/api/products/search')
@login_required
def search_products():
    """Search QuickBooks/Xero products with in-memory caching."""
    import time

    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify({'success': True, 'products': []})

    user_id = current_user.id
    now = time.time()

    cached = _product_cache.get(user_id)
    if cached and (now - cached['cached_at']) < _CACHE_TTL_SECONDS:
        all_products = cached['products']
    else:
        all_products = _load_all_products(user_id)
        _product_cache[user_id] = {'products': all_products, 'cached_at': now}
        current_app.logger.info(f"Product cache refreshed for user {user_id}: {len(all_products)} items")

    q_upper = query.upper()
    results = []
    for p in all_products:
        if (q_upper in (p.get('sku') or '').upper() or
            q_upper in (p.get('name') or '').upper() or
            q_upper in (p.get('description') or '').upper()):
            results.append(p)
            if len(results) >= 25:
                break

    return jsonify({'success': True, 'products': results})


@bp.route('/api/products/cache/clear', methods=['POST'])
@login_required
def clear_product_cache():
    """Manually clear the product cache for current user."""
    _product_cache.pop(current_user.id, None)
    return jsonify({'success': True, 'message': 'Cache cleared'})


def _load_all_products(user_id):
    """Load all products from QuickBooks or Xero for a user."""
    products = []

    try:
        from app.models.quickbooks import QuickBooksConnection
        from app.integrations.quickbooks_service import QuickBooksService

        qb_connection = QuickBooksConnection.query.filter_by(user_id=user_id, is_active=True).first()
        if qb_connection:
            qb_service = QuickBooksService()
            response = qb_service.get_items(qb_connection)
            items = response.get('QueryResponse', {}).get('Item', [])
            for item in items:
                if not item.get('Active', True):
                    continue
                products.append({
                    'id': item.get('Id'),
                    'sku': item.get('Sku', ''),
                    'name': item.get('Name', ''),
                    'description': item.get('Description', ''),
                    'purchase_cost': float(item.get('PurchaseCost', 0) or 0),
                    'unit_price': float(item.get('UnitPrice', 0) or 0),
                    'source': 'quickbooks',
                })
            if products:
                return products
    except Exception as e:
        current_app.logger.warning(f"QB product load error: {e}")

    try:
        from app.models.xero import XeroConnection
        from app.integrations.xero_service import XeroService

        xero_connection = XeroConnection.query.filter_by(user_id=user_id, is_active=True).first()
        if xero_connection:
            xero_service = XeroService()
            items = xero_service.get_items(xero_connection)
            if items:
                for item in items:
                    purchase_price = 0
                    sale_price = 0
                    if item.get('PurchaseDetails'):
                        purchase_price = float(item['PurchaseDetails'].get('UnitPrice', 0) or 0)
                    if item.get('SalesDetails'):
                        sale_price = float(item['SalesDetails'].get('UnitPrice', 0) or 0)
                    products.append({
                        'id': item.get('ItemID'),
                        'sku': item.get('Code', ''),
                        'name': item.get('Name', ''),
                        'description': item.get('Description', ''),
                        'purchase_cost': purchase_price,
                        'unit_price': sale_price,
                        'source': 'xero',
                    })
    except Exception as e:
        current_app.logger.warning(f"Xero product load error: {e}")

    return products


# =============================================================================
# TAKEOFF - LINK PRODUCT TO SYMBOL TYPE
# =============================================================================

@bp.route('/api/projects/<int:project_id>/link-product', methods=['POST'])
@login_required
def link_product_to_symbol(project_id):
    """Link a QB/Xero product to a symbol type template."""
    from app.models.project import Project, ProjectMaterial
    from app.models.takeoff import TakeoffSymbolTemplate, TakeoffSymbolDetection, TakeoffRoom
    from app.extensions import db

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    data = request.get_json()
    template_id = data.get('template_id')
    product = data.get('product')

    if not template_id or not product:
        return jsonify({'success': False, 'error': 'Template and product required'}), 400

    template = TakeoffSymbolTemplate.query.filter_by(
        id=template_id, project_id=project_id
    ).first()
    if not template:
        return jsonify({'success': False, 'error': 'Template not found'}), 404

    template.default_part_number = product.get('sku')
    template.default_product_description = product.get('description') or product.get('name')
    template.default_unit_cost = product.get('purchase_cost')
    template.default_unit_sell = product.get('unit_price')
    template.qb_item_id = product.get('id')

    detections = TakeoffSymbolDetection.query.filter_by(
        project_id=project_id,
        symbol_type_id=template.symbol_type_id,
        rejected=False,
    ).all()

    for det in detections:
        det.part_number = product.get('sku')
        det.product_description = product.get('description') or product.get('name')

    room_counts = {}
    for det in detections:
        room_key = det.room_id or 'unassigned'
        room_counts[room_key] = room_counts.get(room_key, 0) + 1

    materials_created = 0
    for room_key, count in room_counts.items():
        room_id = room_key if room_key != 'unassigned' else None

        if room_id:
            room = TakeoffRoom.query.get(room_id)
            room_cat = f"{room.name} - {template.label}" if room else template.label
        else:
            room_cat = template.label

        existing = ProjectMaterial.query.filter_by(
            project_id=project_id,
            part_number=product.get('sku'),
            category=room_cat,
        ).first()

        if existing:
            existing.quantity = count
            existing.unit_cost = product.get('purchase_cost')
            existing.calculate_totals(markup_percent=float(project.materials_markup_percent))
            qb_sell = product.get('unit_price', 0)
            if qb_sell and qb_sell > float(existing.unit_sell or 0):
                existing.unit_sell = qb_sell
                existing.total_sell = round(count * qb_sell, 2)
        else:
            material = ProjectMaterial(
                project_id=project_id,
                category=room_cat,
                part_number=product.get('sku'),
                description=product.get('description') or product.get('name'),
                quantity=count,
                unit='each',
                unit_cost=product.get('purchase_cost'),
                price_source='quickbooks' if product.get('id') else 'manual',
                price_verified=True if product.get('id') else False,
                qb_item_id=product.get('id'),
                qb_item_name=product.get('name'),
            )
            material.calculate_totals(markup_percent=float(project.materials_markup_percent))
            qb_sell = product.get('unit_price', 0)
            if qb_sell and qb_sell > float(material.unit_sell or 0):
                material.unit_sell = qb_sell
                material.total_sell = round(count * qb_sell, 2)
            db.session.add(material)
            materials_created += 1

    project.recalculate_totals()
    db.session.commit()

    return jsonify({
        'success': True,
        'template': template.to_dict(),
        'materials_created': materials_created,
        'room_counts': {str(k): v for k, v in room_counts.items()},
    })


# =============================================================================
# TAKEOFF - FULL STATE
# =============================================================================

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/takeoff-state')
@login_required
def get_takeoff_state(project_id, doc_id):
    """Get the complete takeoff state for a document."""
    from app.models.project import Project, ProjectDocument
    from app.models.takeoff import (
        TakeoffRoom, TakeoffSymbolDetection, TakeoffSymbolTemplate,
        TakeoffCableRun, TakeoffArea
    )

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    document = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not document:
        return jsonify({'success': False, 'error': 'Document not found'}), 404

    rooms = TakeoffRoom.query.filter_by(project_id=project_id, document_id=doc_id).order_by(TakeoffRoom.sort_order).all()
    templates = TakeoffSymbolTemplate.query.filter_by(project_id=project_id, document_id=doc_id).all()
    detections = TakeoffSymbolDetection.query.filter_by(project_id=project_id, document_id=doc_id, rejected=False).all()
    cable_runs = TakeoffCableRun.query.filter_by(project_id=project_id, document_id=doc_id).all()
    areas = TakeoffArea.query.filter_by(project_id=project_id, document_id=doc_id).all()

    # Accessories - wrapped in try/except so it never breaks the page
    accessories = []
    try:
        from app.models.project import ProjectMaterial
        accessory_materials = ProjectMaterial.query.filter(
            ProjectMaterial.project_id == project_id,
            ProjectMaterial.category.like('%- Accessories')
        ).all()

        for am in accessory_materials:
            parent_label = am.category.replace(' - Accessories', '')
            tpl = next((t for t in templates if t.label == parent_label), None)
            accessories.append({
                'id': am.id,
                'template_id': tpl.id if tpl else None,
                'symbol_type_id': tpl.symbol_type_id if tpl else None,
                'part_number': am.part_number,
                'description': am.description,
                'quantity': am.quantity,
                'unit_cost': float(am.unit_cost or 0),
                'unit_sell': float(am.unit_sell or 0),
            })
    except Exception as e:
        current_app.logger.warning(f"Failed to load accessories: {e}")

    scale = _get_scale(document)

    return jsonify({
        'success': True,
        'document': document.to_dict(),
        'scale': scale,
        'rooms': [r.to_dict() for r in rooms],
        'symbol_templates': [t.to_dict() for t in templates],
        'detections': [d.to_dict() for d in detections],
        'cable_runs': [r.to_dict() for r in cable_runs],
        'areas': [a.to_dict() for a in areas],
        'accessories': accessories,
        'summary': {
            'total_rooms': len(rooms),
            'total_detections': len(detections),
            'total_cable_runs': len(cable_runs),
            'total_areas': len(areas),
        }
    })

# ═══════════════════════════════════════════════════════════════════════════════
# TAKEOFF V8 STATE ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/takeoff-v8-state')
@login_required
def get_takeoff_v8_state(project_id, doc_id):
    """Load v8 takeoff canvas state."""
    from app.models.project import Project, ProjectDocument

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'error': 'Not found'}), 404

    doc = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not doc:
        return jsonify({'error': 'Document not found'}), 404

    state = doc.takeoff_v8_state
    if state:
        return jsonify(state)
    else:
        return jsonify({'empty': True})


@bp.route('/api/projects/<int:project_id>/documents/<int:doc_id>/takeoff-v8-state', methods=['POST'])
@login_required
def save_takeoff_v8_state(project_id, doc_id):
    """Save v8 takeoff canvas state."""
    from app.models.project import Project, ProjectDocument

    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
    if not project:
        return jsonify({'error': 'Not found'}), 404

    doc = ProjectDocument.query.filter_by(id=doc_id, project_id=project_id).first()
    if not doc:
        return jsonify({'error': 'Document not found'}), 404

    state = request.get_json()
    if not state:
        return jsonify({'error': 'No data provided'}), 400

    doc.takeoff_v8_state = state
    db.session.commit()

    return jsonify({'success': True})    
